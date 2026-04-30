from flask import Flask, request, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_cors import CORS
from datetime import datetime, timedelta
import urllib.parse
from flask_migrate import Migrate
from functools import wraps
import jwt
import random
import os
from dotenv import load_dotenv
import docx

load_dotenv()  # 自动读取当前文件夹下的 .env 文件

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.permanent_session_lifetime = timedelta(days=7)

# 确保上传目录存在
import os

# 获取当前文件所在目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 确保上传目录存在
UPLOAD_FOLDER_COMPANY = os.path.join(BASE_DIR, 'uploads', 'company')
os.makedirs(UPLOAD_FOLDER_COMPANY, exist_ok=True)
#日志上传目录
UPLOAD_FOLDER_LOGS = os.path.join(BASE_DIR, 'uploads', 'logs')
os.makedirs(UPLOAD_FOLDER_LOGS, exist_ok=True)

# 简历上传目录
UPLOAD_FOLDER_RESUME = os.path.join(BASE_DIR, 'uploads', 'resume')
os.makedirs(UPLOAD_FOLDER_RESUME, exist_ok=True)

# 附件上传目录
UPLOAD_FOLDER_ATTACHMENTS = os.path.join(BASE_DIR, 'uploads', 'attachments')
os.makedirs(UPLOAD_FOLDER_ATTACHMENTS, exist_ok=True)

# 实验报告上传目录
UPLOAD_FOLDER_EXPERIMENT_REPORTS = os.path.join(BASE_DIR, 'uploads', 'experiment_reports')
os.makedirs(UPLOAD_FOLDER_EXPERIMENT_REPORTS, exist_ok=True)
# JWT 配置
JWT_SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
JWT_EXPIRATION_DELTA = timedelta(days=7)

# 允许的源地址，支持环境变量配置
ALLOWED_ORIGINS = os.environ.get('ALLOWED_ORIGINS', 'http://localhost,http://localhost:5173,http://localhost:80').split(',')

CORS(app, supports_credentials=True, origins=ALLOWED_ORIGINS)

# 确保CORS头被正确设置
@app.after_request
def after_request(response):
    origin = request.headers.get('Origin')
    if origin in ALLOWED_ORIGINS:
        response.headers['Access-Control-Allow-Origin'] = origin
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    return response

# 配置数据库连接信息，使用环境变量
DATABASE_HOST = os.environ.get('DATABASE_HOST', 'localhost')  # 默认使用 localhost
DATABASE_PORT = os.environ.get('DATABASE_PORT', '3306')
DATABASE_USER = os.environ.get('DATABASE_USER', 'root')
DATABASE_PASSWORD = os.environ.get('DATABASE_PASSWORD')
DATABASE_NAME = os.environ.get('DATABASE_NAME', 'bachelor')

# 确保数据库密码存在
if not DATABASE_PASSWORD:
    raise ValueError("DATABASE_PASSWORD environment variable is not set")

app.config['SQLALCHEMY_DATABASE_URI'] = f'mysql+pymysql://{DATABASE_USER}:{DATABASE_PASSWORD}@{DATABASE_HOST}:{DATABASE_PORT}/{DATABASE_NAME}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
migrate = Migrate(app, db)


class DoubleSelectionStep(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    current_step = db.Column(db.Integer, nullable=False, default=0)


class DoubleSelectionTime(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    step = db.Column(db.Integer, nullable=False)  # 0: 准备阶段, 1: 学生选导师, 2: 导师确认, 3: 双选完成
    start_time = db.Column(db.DateTime, nullable=True)  # 开始时间
    end_time = db.Column(db.DateTime, nullable=True)  # 结束时间
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('step', name='_step_uc'),
    )



# 定义用户模型
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    username = db.Column(db.String(80), unique=True, nullable=False)  # 学号/工号
    password_hash = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # student/teacher/admin
    realname = db.Column(db.String(20), nullable=False)
    id_card_last6 = db.Column(db.String(6), nullable=True)  # 身份证后六位（首次登录用）
    first_login = db.Column(db.Boolean, default=True)  # 是否首次登录
    
    major = db.Column(db.String(255), nullable=True)
    department = db.Column(db.String(255), nullable=True)  # 院系
    contact = db.Column(db.String(255), nullable=True)  # 联系方式
    title = db.Column(db.String(100), nullable=True)  # 职称
    email = db.Column(db.String(255), nullable=True)  # 电子邮箱
    is_confirmed = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)  # 账号是否激活（用于企业用户）

# 校外企业用户模型
class Company(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)  # 关联的用户ID
    company_name = db.Column(db.String(255), nullable=False)  # 企业名称
    field = db.Column(db.String(255), nullable=False)  # 领域
    nature = db.Column(db.String(100), nullable=False)  # 性质（国企、私企、外企等）
    scale = db.Column(db.String(100), nullable=False)  # 规模（小型、中型、大型等）


    contact = db.Column(db.String(255), nullable=False)  # 联系方式
    proof_file = db.Column(db.String(255), nullable=False)  # 证明材料文件名
    status = db.Column(db.String(20), nullable=False, default='pending')  # pending/approved/rejected
    description = db.Column(db.Text, nullable=True)  # 企业描述
    create_time = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    review_time = db.Column(db.DateTime, nullable=True)
    review_comment = db.Column(db.Text, nullable=True)
    
    # 关系
    user = db.relationship('User', backref='company', uselist=False)


# 双选导师表
class DoubleSelectionTeacher(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)  # 关联的教师ID
    min_quota = db.Column(db.Integer, nullable=True)  # 最小名额
    max_quota = db.Column(db.Integer, nullable=True)  # 最大名额
    current_quota = db.Column(db.Integer, nullable=True, default=0)  # 当前名额
    status = db.Column(db.String(20), nullable=False, default='active')  # 状态：active/inactive
    join_time = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)  # 加入时间
    
    # 关联关系
    teacher = db.relationship('User', backref=db.backref('double_selection', uselist=False))

class DoubleSelectionStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)  # 关联的学生ID
    status = db.Column(db.String(20), nullable=False, default='active')  # 状态：active/inactive
    join_time = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)  # 加入时间
    
    # 关联关系
    student = db.relationship('User', backref=db.backref('double_selection_student', uselist=False))

# 实验报告表
class ExperimentReport(db.Model):
    __tablename__ = 'experiment_report'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    project_id = db.Column(db.Integer, db.ForeignKey('research_project.id'), nullable=False)  # 绑定的科研项目
    title = db.Column(db.String(100), nullable=False)
    content = db.Column(db.Text, nullable=False)
    date = db.Column(db.Date, nullable=False)
    file_url = db.Column(db.String(255), nullable=True)
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 关联
    student = db.relationship('User', backref=db.backref('experiment_reports', lazy=True))
    project = db.relationship('ResearchProject', backref=db.backref('experiment_reports', lazy=True))
# 实习招聘表
class Internship(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    company_id = db.Column(db.Integer, db.ForeignKey('company.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    position = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text, nullable=False)
    requirements = db.Column(db.Text, nullable=False)
    city = db.Column(db.String(100), nullable=False)  # 城市
    location = db.Column(db.String(255), nullable=False)  # 具体地点
    salary = db.Column(db.String(100))
    skill_tags = db.Column(db.Text)
    welfare_tags = db.Column(db.Text)
    quota = db.Column(db.Integer, nullable=False)
    deadline = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='active')  # active, closed
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # 关系
    company = db.relationship('Company', backref=db.backref('internships', lazy='dynamic'))

# 工作日志模型
class WorkLog(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 学生ID
    content = db.Column(db.Text, nullable=False)  # 日志内容
    date = db.Column(db.Date, nullable=False)  # 日志日期
    file = db.Column(db.String(255), nullable=True)  # 附件文件名
    created_at = db.Column(db.DateTime, default=datetime.utcnow)  # 创建时间
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)  # 更新时间
    
    # 关系
    student = db.relationship('User', backref=db.backref('work_logs', lazy='dynamic'))

# 实习申请表
class InternshipApplication(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    internship_id = db.Column(db.Integer, db.ForeignKey('internship.id'), nullable=False)
    resume_file = db.Column(db.String(255), nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending, approved, rejected
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # 关系
    student = db.relationship('User', backref=db.backref('internship_applications', lazy='dynamic'))
    internship = db.relationship('Internship', backref=db.backref('applications', lazy='dynamic'))

# 导师双选记录模型
class TeacherSelection(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    priority = db.Column(db.Integer, nullable=False)  # 1: 第一志愿, 2: 第二志愿
    status = db.Column(db.String(20), nullable=False, default='pending')  # pending/confirmed/rejected
    create_time = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    
    # 关系
    student = db.relationship('User', foreign_keys=[student_id], backref='selections_as_student')
    teacher = db.relationship('User', foreign_keys=[teacher_id], backref='selections_as_teacher')
    
    # 唯一约束：每个学生每个志愿只能选择一个导师
    __table_args__ = (
        db.UniqueConstraint('student_id', 'priority', name='_student_priority_uc'),
    )
    
    # 唯一约束：每个学生每个志愿只能选择一个导师


# 双选结果表
class DoubleSelectionResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    status = db.Column(db.String(20), default='confirmed')
    create_time = db.Column(db.DateTime, default=datetime.utcnow)

# 科研训练项目模型
class ResearchProject(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    batch = db.Column(db.String(20), nullable=False)  # 批次，如2025-2026学年春季
    project_name = db.Column(db.String(255), nullable=False)  # 项目名称
    direction = db.Column(db.String(255), nullable=True)  # 研究方向
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 导师ID
    max_students = db.Column(db.Integer, nullable=True)  # 招收学生人数
    registered_count = db.Column(db.Integer, default=0)  # 已报名人数
    confirmed_count = db.Column(db.Integer, default=0)  # 已确认人数
    max_classes = db.Column(db.Integer, nullable=True)  # 可报名最大班级数
    attendance_requirement = db.Column(db.String(255), nullable=True)  # 点名要求
    major_ranking = db.Column(db.String(50), nullable=True)  # 专业排名要求
    allow_failed = db.Column(db.Boolean, default=False)  # 是否允许不及格成绩
    major = db.Column(db.String(255), nullable=True)  # 所属专业
    department = db.Column(db.String(255), nullable=True)  # 所属院系
    is_national = db.Column(db.Boolean, default=False)  # 是否属于国家/教育部项目
    national_lab = db.Column(db.String(255), nullable=True)  # 所属国家（教育部）重点实验室
    description = db.Column(db.Text, nullable=True)  # 课题简介
    requirements = db.Column(db.Text, nullable=True)  # 报名要求
    status = db.Column(db.String(20), default='pending')  # pending/approved/rejected
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 关系
    teacher = db.relationship('User', backref='research_projects')

# 科研训练报名模型
class ResearchRegistration(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 学生ID
    project_id = db.Column(db.Integer, db.ForeignKey('research_project.id'), nullable=False)  # 项目ID
    status = db.Column(db.String(20), default='pending')  # pending/confirmed/rejected
    create_time = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 关系
    student = db.relationship('User', backref='research_registrations')
    project = db.relationship('ResearchProject', backref='registrations')
    
    # 唯一约束：每个学生每个项目只能报名一次
    __table_args__ = (
        db.UniqueConstraint('student_id', 'project_id', name='_student_project_uc'),
    )

# 公告模型
class Announcement(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    title = db.Column(db.String(255), nullable=False)  # 公告标题
    content = db.Column(db.Text, nullable=False)  # 公告内容
    author_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 发布者ID
    create_time = db.Column(db.DateTime, default=datetime.utcnow)  # 发布时间
    
    # 关系
    author = db.relationship('User', backref='announcements')
    attachments = db.relationship('Attachment', backref='announcement', cascade='all, delete-orphan')

# 附件模型
class Attachment(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    filename = db.Column(db.String(255), nullable=False)  # 文件名
    filepath = db.Column(db.String(500), nullable=False)  # 文件路径
    size = db.Column(db.Integer, nullable=False)  # 文件大小（KB）
    announcement_id = db.Column(db.Integer, db.ForeignKey('announcement.id'), nullable=False)  # 所属公告ID
    upload_time = db.Column(db.DateTime, default=datetime.utcnow)  # 上传时间

# 团队模型
class Team(db.Model):
    __tablename__ = 'team'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    theme = db.Column(db.String(255), nullable=False)  # 团队主题
    course = db.Column(db.String(255), nullable=False)  # 课程名称
    description = db.Column(db.Text, nullable=True)  # 团队描述
    status = db.Column(db.String(20), nullable=False, default='pending')  # 状态：pending/approved/rejected/dissolving/dissolved
    initiator_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 发起者ID
    create_time = db.Column(db.DateTime, default=datetime.utcnow)  # 创建时间
    review_time = db.Column(db.DateTime, nullable=True)  # 审核时间
    review_comment = db.Column(db.Text, nullable=True)  # 审核意见
    dissolve_reason = db.Column(db.Text, nullable=True)  # 解散原因
    dissolve_time = db.Column(db.DateTime, nullable=True)  # 解散申请时间
    
    # 关联
    initiator = db.relationship('User', backref=db.backref('initiated_teams', lazy=True))

# 团队成员模型
class TeamMember(db.Model):
    __tablename__ = 'team_member'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)  # 团队ID
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 用户ID
    status = db.Column(db.String(20), nullable=False, default='pending')  # 状态：pending/accepted/rejected
    join_time = db.Column(db.DateTime, default=datetime.utcnow)  # 加入时间
    
    # 关联
    team = db.relationship('Team', backref=db.backref('members', lazy=True))
    user = db.relationship('User', backref=db.backref('team_memberships', lazy=True))
    
    # 唯一约束：每个用户在每个团队中只能有一条记录
    __table_args__ = (db.UniqueConstraint('team_id', 'user_id', name='_team_user_uc'),)

# 申请日志模型
class ApplicationLog(db.Model):
    __tablename__ = 'application_log'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 用户ID
    type = db.Column(db.String(50), nullable=False)  # 申请类型：internship/train/double-selection/team/company/project
    status = db.Column(db.String(20), nullable=False)  # 申请状态：pending/approved/rejected/completed
    handler_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # 处理人ID
    handler_name = db.Column(db.String(255), nullable=True)  # 处理人姓名
    receipt = db.Column(db.Text, nullable=True)  # 回执信息
    description = db.Column(db.Text, nullable=True)  # 详细描述
    create_time = db.Column(db.DateTime, default=datetime.utcnow)  # 创建时间
    
    # 关联
    user = db.relationship('User', foreign_keys=[user_id], backref='application_logs')
    handler = db.relationship('User', foreign_keys=[handler_id])




# 公告阅读记录模型
class AnnouncementReadRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # 用户ID
    announcement_id = db.Column(db.Integer, db.ForeignKey('announcement.id'), nullable=False)  # 公告ID
    read_time = db.Column(db.DateTime, default=datetime.utcnow)  # 阅读时间
    
    # 唯一约束，确保每个用户对每个公告只有一条阅读记录
    __table_args__ = (db.UniqueConstraint('user_id', 'announcement_id', name='_user_announcement_uc'),)

# JWT Token 工具函数
def generate_token(user_id, username, role):
    """生成 JWT Token"""
    payload = {
        'user_id': user_id,
        'username': username,
        'role': role,
        'is_company': role == 'company',
        'exp': datetime.utcnow() + JWT_EXPIRATION_DELTA,
        'iat': datetime.utcnow()
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm='HS256')

def decode_token(token):
    """解码 JWT Token"""
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        return None  # Token 过期
    except jwt.InvalidTokenError as e:
        print(f"Token decode error: {str(e)}")  # 打印错误信息
        return None  # Token 无效

def get_token_from_header():
    """从请求头获取 Token，同时检查URL查询参数"""
    # 1. 先从请求头获取
    auth_header = request.headers.get('Authorization')
    if auth_header and auth_header.startswith('Bearer '):
        return auth_header.split(' ')[1]
    
    # 2. 如果请求头中没有，从URL查询参数获取
    token = request.args.get('token')
    if token:
        return token
    
    return None

# 新的 JWT 登录装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = get_token_from_header()
        
        if not token:
            return jsonify({'success': False, 'message': '缺少认证令牌'}), 401
        
        payload = decode_token(token)
        if not payload:
            return jsonify({'success': False, 'message': '令牌无效或已过期'}), 401
        
        # 检查payload中是否包含必要的字段
        if 'user_id' not in payload:
            return jsonify({'success': False, 'message': '令牌格式无效'}), 401
        
        # 将用户信息存入请求上下文
        request.current_user = payload
        return f(*args, **kwargs)
    return decorated_function

def get_teacher_quota(teacher_id):
    """获取教师的名额信息"""
    # 从双选导师表中获取名额信息
    double_teacher = DoubleSelectionTeacher.query.filter_by(teacher_id=teacher_id).first()
    if not double_teacher:
        return {'current': 0, 'max': 5}  # 默认最大名额为5
    return {'current': double_teacher.current_quota or 0, 'max': double_teacher.max_quota or 5}

def update_teacher_quota(teacher_id, increment=True):
    """更新教师的名额信息"""
    double_teacher = DoubleSelectionTeacher.query.filter_by(teacher_id=teacher_id).first()
    if double_teacher:
        if increment:
            double_teacher.current_quota = (double_teacher.current_quota or 0) + 1
        else:
            double_teacher.current_quota = max(0, (double_teacher.current_quota or 1) - 1)
        db.session.commit()

# 注册API
@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    realname = data.get('realname')
    role = data.get('role', 'student')
    major = data.get('major')
    
    if not username or not password or not realname:
        return jsonify({'success': False, 'message': '用户名、密码和真实姓名不能为空'}), 400
    
    # 检查用户名是否已存在
    existing_user = User.query.filter_by(username=username).first()
    if existing_user:
        return jsonify({'success': False, 'message': '用户名已存在'}), 400
    
    # 密码加密
    password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    
    # 创建新用户
    new_user = User(
        username=username,
        password_hash=password_hash,
        realname=realname,
        role=role,
        major=major
    )
    
    try:
        db.session.add(new_user)
        db.session.commit()
        return jsonify({'success': True, 'message': '注册成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'注册失败: {str(e)}'}), 500



# 校外企业注册API
@app.route('/api/company-register', methods=['POST'])
def company_register():
    if 'proof' not in request.files:
        return jsonify({'success': False, 'message': '请上传证明材料'}), 400
    
    proof_file = request.files['proof']
    if proof_file.filename == '':
        return jsonify({'success': False, 'message': '请选择证明材料文件'}), 400
    
    username = request.form.get('username')
    password = request.form.get('password')
    realname = request.form.get('realname')
    company_name = request.form.get('company_name')
    field = request.form.get('field')
    nature = request.form.get('nature')
    scale = request.form.get('scale')
    contact = request.form.get('contact')
    
    if not all([username, password, realname, company_name, field, nature, scale, contact]):
        return jsonify({'success': False, 'message': '请填写完整信息'}), 400
    
    # 检查用户名是否已存在
    existing_user = User.query.filter_by(username=username).first()
    if existing_user:
        return jsonify({'success': False, 'message': '用户名已存在'}), 400
    
    try:
        # 保存证明材料文件
        import uuid
        filename = f"{uuid.uuid4()}_{proof_file.filename}"
        filepath = os.path.join(UPLOAD_FOLDER_COMPANY, filename)
        proof_file.save(filepath)
        
        # 创建用户（企业用户）
        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        new_user = User(
            username=username,
            password_hash=password_hash,
            realname=realname,
            role='company',
            is_active=False,  # 默认未激活，需管理员审批
            first_login=False  # 企业用户不需要首次登录改密码
        )
        
        db.session.add(new_user)
        db.session.flush()  # 获取user_id
        
        # 创建企业记录
        new_company = Company(
            user_id=new_user.id,
            company_name=company_name,
            field=field,
            nature=nature,
            scale=scale,
            contact=contact,
            proof_file=filename,
            status='pending'
        )
        
        db.session.add(new_company)
        db.session.commit()
        
        # 添加申请日志
        log = ApplicationLog(
            user_id=new_user.id,
            type='company',
            status='pending',
            receipt='企业注册申请',
            description=f'企业注册申请：{company_name}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '注册成功，请等待管理员审批'})
    except Exception as e:
        db.session.rollback()
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'success': False, 'message': f'注册失败: {str(e)}'}), 500

# 获取企业信息
@app.route('/api/company/info', methods=['GET'])
@login_required
def get_company_info():
    """获取企业信息"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        if user.role != 'company':
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        company = Company.query.filter_by(user_id=user_id).first()
        
        if not company:
            return jsonify({'success': False, 'message': '企业信息不存在'})
        
        return jsonify({
            'success': True,
            'company': {
                'id': company.id,
                'company_name': company.company_name,
                'field': company.field,
                'nature': company.nature,
                'scale': company.scale,
                'contact': company.contact,
                'proof_file': company.proof_file,
                'status': company.status,
                'description': company.description,
                'review_comment': company.review_comment,
                'create_time': company.create_time.strftime('%Y-%m-%d %H:%M:%S') if company.create_time else None,
                'review_time': company.review_time.strftime('%Y-%m-%d %H:%M:%S') if company.review_time else None
            }
        })
    except Exception as e:
        print(f"获取企业信息失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取企业信息失败: {str(e)}'})

# 重新上传证明材料
@app.route('/api/company/reupload-proof', methods=['POST'])
@login_required
def reupload_proof():
    """重新上传证明材料"""
    try:
        # 检查请求是否包含文件
        if 'proof_file' not in request.files:
            return jsonify({'success': False, 'message': '请上传证明材料'})
        
        file = request.files['proof_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '请选择文件'})
        
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        if user.role != 'company':
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        company = Company.query.filter_by(user_id=user_id).first()
        
        if not company:
            return jsonify({'success': False, 'message': '企业信息不存在'})
        
        # 生成唯一文件名
        import uuid
        filename = f"{uuid.uuid4()}_{file.filename}"
        filepath = os.path.join(UPLOAD_FOLDER_COMPANY, filename)
        file.save(filepath)
        
        # 删除旧文件
        old_filepath = os.path.join(UPLOAD_FOLDER_COMPANY, company.proof_file)
        if os.path.exists(old_filepath):
            os.remove(old_filepath)
        
        # 更新企业信息
        company.proof_file = filename
        company.status = 'pending'  # 重新设置为待审核状态
        company.review_comment = None
        company.review_time = None
        
        # 更新用户状态为未激活
        user.is_active = False
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': '证明材料已重新提交，等待审核'})
    except Exception as e:
        db.session.rollback()
        # 清理上传的文件
        if 'filepath' in locals() and os.path.exists(filepath):
            os.remove(filepath)
        print(f"重新上传证明材料失败: {str(e)}")
        return jsonify({'success': False, 'message': f'重新上传失败: {str(e)}'})

# 企业发布实习招聘
@app.route('/api/company/publish-internship', methods=['POST'])
@login_required
def publish_internship():
    """企业发布实习招聘"""
    try:
        data = request.get_json()
        user_id = request.current_user['user_id']
        
        # 验证企业信息
        company = Company.query.filter_by(user_id=user_id).first()
        if not company:
            return jsonify({'success': False, 'message': '企业信息不存在'})
        
        # 验证必填字段
        required_fields = ['title', 'position', 'description', 'requirements', 'city', 'location', 'quota', 'deadline']
        for field in required_fields:
            if field not in data:
                return jsonify({'success': False, 'message': f'缺少必填字段: {field}'})
        
        # 创建实习招聘
        internship = Internship(
            company_id=company.id,
            title=data['title'],
            position=data['position'],
            description=data['description'],
            requirements=data['requirements'],
            city=data['city'],
            location=data['location'],
            salary=data.get('salary', ''),
            skill_tags=','.join(data.get('skill_tags', [])),
            welfare_tags=','.join(data.get('welfare_tags', [])),
            quota=data['quota'],
            deadline=data['deadline'],
            status='active'
        )
        
        db.session.add(internship)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '实习招聘发布成功'})
    except Exception as e:
        db.session.rollback()
        print(f"发布实习招聘失败: {str(e)}")
        return jsonify({'success': False, 'message': f'发布失败: {str(e)}'})

# 学生申请实习
@app.route('/api/internship/apply', methods=['POST'])
@login_required
def apply_internship():
    """学生申请实习"""
    try:
        # 检查请求是否包含文件
        if 'resume' not in request.files:
            return jsonify({'success': False, 'message': '请上传简历'})
        
        file = request.files['resume']
        if file.filename == '':
            return jsonify({'success': False, 'message': '请选择文件'})
        
        # 验证文件类型（只允许pdf和docx）
        allowed_extensions = {'pdf', 'docx'}
        file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else ''
        
        if file_extension not in allowed_extensions:
            return jsonify({'success': False, 'message': '请上传PDF或DOCX格式的简历文件，不支持DOC格式'})
        
        internship_id = request.form.get('internship_id')
        if not internship_id:
            return jsonify({'success': False, 'message': '缺少实习ID'})
        
        user_id = request.current_user['user_id']
        
        # 检查实习是否存在
        internship = Internship.query.get(internship_id)
        if not internship:
            return jsonify({'success': False, 'message': '实习信息不存在'})
        
        # 检查学生是否已有确认的实习记录
        confirmed_application = InternshipApplication.query.filter_by(
            student_id=user_id,
            status='confirmed'
        ).first()
        if confirmed_application:
            return jsonify({'success': False, 'message': '您已有一份已确认的实习记录，不能再申请其他实习'})
        
        # 检查是否已经申请
        existing_application = InternshipApplication.query.filter_by(
            student_id=user_id,
            internship_id=internship_id
        ).first()
        if existing_application:
            return jsonify({'success': False, 'message': '您已经申请过该实习'})
        
        # 生成唯一文件名
        import uuid
        # 清理文件名，移除特殊字符
        import re
        clean_filename = re.sub(r'[<>:"/\\|?*]', '_', file.filename)
        filename = f"{uuid.uuid4()}_{clean_filename}"
        filepath = os.path.join(UPLOAD_FOLDER_RESUME, filename)
        file.save(filepath)
        print(f"文件已保存到: {filepath}")
        
        # 创建申请
        application = InternshipApplication(
            student_id=user_id,
            internship_id=internship_id,
            resume_file=filename,
            status='pending'
        )
        
        db.session.add(application)
        db.session.commit()
        
        # 添加申请日志
        log = ApplicationLog(
            user_id=user_id,
            type='internship',
            status='pending',
            receipt='申请实习',
            description=f'申请实习：{internship.title}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '申请成功'})
    except Exception as e:
        db.session.rollback()
        # 清理上传的文件
        if 'filepath' in locals() and os.path.exists(filepath):
            os.remove(filepath)
        print(f"申请实习失败: {str(e)}")
        return jsonify({'success': False, 'message': f'申请失败: {str(e)}'})

# 企业获取学生申请列表
@app.route('/api/company/applications', methods=['GET'])
@login_required
def get_company_applications():
    """企业获取学生申请列表"""
    try:
        user_id = request.current_user['user_id']
        
        # 获取企业信息
        company = Company.query.filter_by(user_id=user_id).first()
        if not company:
            return jsonify({'success': False, 'message': '企业信息不存在'})
        
        # 获取企业发布的实习
        internships = Internship.query.filter_by(company_id=company.id).all()
        internship_ids = [internship.id for internship in internships]
        
        # 获取申请（只获取状态为pending的申请，用于实习审核）
        applications = InternshipApplication.query.filter(
            InternshipApplication.internship_id.in_(internship_ids),
            InternshipApplication.status == 'pending'
        ).all()
        
        # 构建响应数据
        application_list = []
        for app in applications:
            student = User.query.get(app.student_id)
            internship = Internship.query.get(app.internship_id)
            application_list.append({
                'id': app.id,
                'student_id': student.username,
                'student_name': student.realname,
                'major': student.major or '',
                'position': internship.position,
                'apply_time': app.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'status': app.status,
                'resume_file': app.resume_file
            })
        
        return jsonify({'success': True, 'applications': application_list})
    except Exception as e:
        print(f"获取申请列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 企业处理申请
@app.route('/api/company/application/process', methods=['POST'])
@login_required
def process_application():
    """企业处理申请"""
    try:
        data = request.get_json()
        application_id = data.get('application_id')
        action = data.get('action')  # approve 或 reject
        
        if not application_id or not action:
            return jsonify({'success': False, 'message': '缺少参数'})
        
        if action not in ['approve', 'reject']:
            return jsonify({'success': False, 'message': '无效的操作'})
        
        user_id = request.current_user['user_id']
        
        # 获取企业信息
        company = Company.query.filter_by(user_id=user_id).first()
        if not company:
            return jsonify({'success': False, 'message': '企业信息不存在'})
        
        # 获取申请
        application = InternshipApplication.query.get(application_id)
        if not application:
            return jsonify({'success': False, 'message': '申请不存在'})
        
        # 检查申请是否属于该企业
        internship = Internship.query.get(application.internship_id)
        if internship.company_id != company.id:
            return jsonify({'success': False, 'message': '权限不足'})
        
        # 更新状态
        application.status = 'approved' if action == 'approve' else 'rejected'
        application.updated_at = datetime.now()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': '处理成功'})
    except Exception as e:
        db.session.rollback()
        print(f"处理申请失败: {str(e)}")
        return jsonify({'success': False, 'message': f'处理失败: {str(e)}'})

@app.route('/api/company/update-description', methods=['POST'])
@login_required
def update_company_description():
    """更新企业描述"""
    try:
        # 检查用户角色
        if request.current_user.get('role') != 'company':
            return jsonify({'success': False, 'message': '权限不足'})
        
        data = request.get_json()
        if not data or 'description' not in data:
            return jsonify({'success': False, 'message': '请提供企业描述'})
        
        description = data['description']
        user_id = request.current_user['user_id']
        
        # 更新企业信息
        company = Company.query.filter_by(user_id=user_id).first()
        if company:
            company.description = description
            db.session.commit()
            return jsonify({'success': True, 'message': '企业描述已更新'})
        else:
            return jsonify({'success': False, 'message': '企业信息不存在'})
    except Exception as e:
        db.session.rollback()
        print(f"更新企业描述失败: {str(e)}")
        return jsonify({'success': False, 'message': '更新失败，请重试'})

@app.route('/api/company/confirmed-students', methods=['GET'])
@login_required
def get_confirmed_students():
    """获取已确认的实习生列表"""
    try:
        # 检查用户角色
        if request.current_user.get('role') != 'company':
            return jsonify({'success': False, 'message': '权限不足'})
        
        user_id = request.current_user['user_id']
        
        # 获取企业信息
        company = Company.query.filter_by(user_id=user_id).first()
        if not company:
            return jsonify({'success': False, 'message': '企业信息不存在'})
        
        # 获取该企业的实习岗位
        internships = Internship.query.filter_by(company_id=company.id).all()
        internship_ids = [internship.id for internship in internships]
        
        # 获取已批准和已确认的申请
        applications = InternshipApplication.query.filter(
            InternshipApplication.internship_id.in_(internship_ids),
            InternshipApplication.status.in_(['approved', 'confirmed'])
        ).all()
        
        # 构建学生列表
        student_list = []
        for app in applications:
            student = User.query.filter_by(id=app.student_id).first()
            internship = Internship.query.filter_by(id=app.internship_id).first()
            
            # 获取学生的导师信息
            teacher = None
            if student:
                # 从双选表中获取导师
                selection = DoubleSelectionStudent.query.filter_by(student_id=student.id).first()
                if selection and selection.teacher_id:
                    teacher = User.query.filter_by(id=selection.teacher_id).first()
            
            if student and internship:
                student_list.append({
                    'id': app.id,
                    'student_id': student.username,
                    'student_name': student.realname,
                    'major': student.major or '',
                    'position': internship.position,
                    'confirm_time': app.updated_at.strftime('%Y-%m-%d %H:%M:%S') if app.updated_at else '',
                    'resume_file': app.resume_file,
                    'student_contact': student.contact or '',
                    'student_email': student.email or '',
                    'teacher_name': teacher.realname if teacher else '',
                    'teacher_contact': teacher.contact if teacher else '',
                    'teacher_email': teacher.email if teacher else ''
                })
        
        return jsonify({'success': True, 'students': student_list})
    except Exception as e:
        print(f"获取已确认学生失败: {str(e)}")
        return jsonify({'success': False, 'message': '获取失败，请重试'})

# 查看简历
@app.route('/api/internship/resume/<filename>', methods=['GET'])
@login_required
def view_resume(filename):
    """查看简历"""
    try:
        # 安全处理文件名，防止路径遍历攻击
        import posixpath
        import ntpath
        # 规范化文件名，移除路径部分
        filename = posixpath.basename(filename)
        filename = ntpath.basename(filename)
        
        # 构建文件路径
        filepath = os.path.join(UPLOAD_FOLDER_RESUME, filename)
        print(f"尝试读取文件: {filepath}")
        
        if not os.path.exists(filepath):
            print(f"文件不存在: {filepath}")
            # 列出目录中的文件，方便调试
            files = os.listdir(UPLOAD_FOLDER_RESUME)
            print(f"目录中的文件: {files}")
            return jsonify({'success': False, 'message': '文件不存在'}), 404
        
        return send_file(filepath)
    except Exception as e:
        print(f"查看简历失败: {str(e)}")
        return jsonify({'success': False, 'message': f'查看失败: {str(e)}'})

# 获取Word文档文本内容（用于预览）
@app.route('/api/internship/resume/text/<filename>', methods=['GET'])
@login_required
def get_resume_text(filename):
    """获取Word文档文本内容"""
    try:
        # 安全处理文件名
        import posixpath
        import ntpath
        filename = posixpath.basename(filename)
        filename = ntpath.basename(filename)
        
        filepath = os.path.join(UPLOAD_FOLDER_RESUME, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404
        
        # 获取文件扩展名
        ext = filename.split('.')[-1].lower()
        
        if ext == 'docx':
            try:
                from docx import Document
                doc = Document(filepath)
                full_text = []
                for para in doc.paragraphs:
                    if para.text.strip():
                        full_text.append(para.text)
                text_content = '\n'.join(full_text)
                return jsonify({'success': True, 'content': text_content})
            except ImportError:
                print("错误：未安装python-docx库")
                return jsonify({'success': False, 'message': '未安装python-docx库'}), 500
            except Exception as e:
                print(f"解析docx失败: {str(e)}")
                return jsonify({'success': False, 'message': f'解析失败: {str(e)}'}), 500
        elif ext == 'doc':
            # .doc格式不支持在线预览，需要下载查看
            print("提示：.doc格式不支持在线预览")
            return jsonify({'success': False, 'message': '.doc格式不支持在线预览，请下载后查看'}), 400
        
        return jsonify({'success': False, 'message': '不支持的文件格式'}), 400
    except Exception as e:
        print(f"获取简历文本失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 检查学生是否已有确认的实习记录
@app.route('/api/student/has-confirmed-internship', methods=['GET'])
@login_required
def check_has_confirmed_internship():
    """检查学生是否已有确认的实习记录"""
    try:
        user_id = request.current_user['user_id']
        
        confirmed_application = InternshipApplication.query.filter_by(
            student_id=user_id,
            status='confirmed'
        ).first()
        
        return jsonify({'success': True, 'has_confirmed': confirmed_application is not None})
    except Exception as e:
        print(f"检查实习记录失败: {str(e)}")
        return jsonify({'success': False, 'message': '检查失败', 'has_confirmed': False})

# 学生获取自己的实习申请列表
@app.route('/api/student/applications', methods=['GET'])
@login_required
def get_student_applications():
    """学生获取自己的实习申请列表"""
    try:
        user_id = request.current_user['user_id']
        
        # 获取学生的所有申请
        applications = InternshipApplication.query.filter_by(student_id=user_id).all()
        
        # 构建响应数据
        application_list = []
        for app in applications:
            internship = Internship.query.get(app.internship_id)
            company = Company.query.get(internship.company_id)
            application_list.append({
                'id': app.id,
                'company': company.company_name,
                'position': internship.position,
                'location': f"{internship.city} {internship.location}",
                'status': app.status,
                'apply_time': app.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'internship_id': internship.id
            })
        
        return jsonify({'success': True, 'applications': application_list})
    except Exception as e:
        print(f"获取申请列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 学生确认实习
@app.route('/api/student/confirm-internship', methods=['POST'])
@login_required
def confirm_internship():
    """学生确认实习，自动拒绝其他所有批准的申请"""
    try:
        data = request.get_json()
        application_id = data.get('application_id')
        
        if not application_id:
            return jsonify({'success': False, 'message': '缺少参数'}), 400
        
        user_id = request.current_user['user_id']
        
        # 获取要确认的申请
        confirm_application = InternshipApplication.query.get(application_id)
        if not confirm_application:
            return jsonify({'success': False, 'message': '申请不存在'}), 404
        
        # 验证申请是否属于该学生
        if confirm_application.student_id != user_id:
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        # 验证申请状态是否为已批准
        if confirm_application.status != 'approved':
            return jsonify({'success': False, 'message': '只能确认已批准的申请'}), 400
        
        # 确认该申请
        confirm_application.status = 'confirmed'
        
        # 自动拒绝其他所有已批准的申请
        other_applications = InternshipApplication.query.filter(
            InternshipApplication.student_id == user_id,
            InternshipApplication.id != application_id,
            InternshipApplication.status == 'approved'
        ).all()
        
        for app in other_applications:
            app.status = '自动拒绝'
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': '确认成功，已自动拒绝其他申请'})
    except Exception as e:
        db.session.rollback()
        print(f"确认实习失败: {str(e)}")
        return jsonify({'success': False, 'message': f'确认失败: {str(e)}'})

@app.route('/api/student/logs', methods=['POST'])
@login_required
def submit_student_log():
    """学生提交工作日志"""
    try:
        # 检查请求是否包含文件
        file = request.files.get('file')
        
        content = request.form.get('content')
        date = request.form.get('date')
        
        if not content or not date:
            return jsonify({'success': False, 'message': '缺少必填字段'}), 400
        
        user_id = request.current_user['user_id']
        
        # 保存文件（如果有）
        filename = None
        if file and file.filename:
            import uuid
            # 清理文件名，移除特殊字符
            import re
            clean_filename = re.sub(r'[<>:"/\\|?*]', '_', file.filename)
            filename = f"{uuid.uuid4()}_{clean_filename}"
            filepath = os.path.join(UPLOAD_FOLDER_LOGS, filename)
            file.save(filepath)
        
        # 创建日志
        log = WorkLog(
            student_id=user_id,
            content=content,
            date=date,
            file=filename
        )
        
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '日志提交成功'})
    except Exception as e:
        db.session.rollback()
        # 清理上传的文件
        if 'filepath' in locals() and os.path.exists(filepath):
            os.remove(filepath)
        print(f"提交日志失败: {str(e)}")
        return jsonify({'success': False, 'message': f'提交失败: {str(e)}'})

# 学生获取自己的工作日志列表
@app.route('/api/student/logs', methods=['GET'])
@login_required
def get_student_logs():
    """学生获取自己的工作日志列表"""
    try:
        user_id = request.current_user['user_id']
        
        # 获取学生的所有日志
        logs = WorkLog.query.filter_by(student_id=user_id).order_by(WorkLog.date.desc()).all()
        
        # 构建响应数据
        log_list = []
        for log in logs:
            log_list.append({
                'id': log.id,
                'content': log.content,
                'date': log.date.strftime('%Y-%m-%d'),
                'file': log.file,
                'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({'success': True, 'logs': log_list})
    except Exception as e:
        print(f"获取日志列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 教师获取学生列表
@app.route('/api/teacher/students', methods=['GET'])
@login_required
def get_teacher_students():
    """教师获取自己的学生列表（支持搜索）"""
    try:
        user_id = request.current_user['user_id']
        keyword = request.args.get('keyword', '').strip()
        
        # 通过双选结果获取学生
        results = DoubleSelectionResult.query.filter_by(teacher_id=user_id).all()
        
        # 构建学生列表
        students = []
        for result in results:
            student = User.query.get(result.student_id)
            if student:
                # 如果有搜索关键词，进行过滤
                if keyword:
                    if keyword.lower() not in student.realname.lower() and keyword.lower() not in student.username.lower():
                        continue
                students.append({
                    'id': student.id,
                    'name': student.realname,
                    'username': student.username,
                    'major': student.major
                })
        
        return jsonify({'success': True, 'students': students})
    except Exception as e:
        print(f"获取学生列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 教师获取学生的工作日志
@app.route('/api/teacher/student-logs/<int:student_id>', methods=['GET'])
@login_required
def get_student_logs_by_teacher(student_id):
    """教师获取学生的工作日志（支持日期筛选）"""
    try:
        teacher_id = request.current_user['user_id']
        
        # 获取日期筛选参数
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # 验证学生是否属于该教师
        result = DoubleSelectionResult.query.filter_by(
            teacher_id=teacher_id,
            student_id=student_id
        ).first()
        
        if not result:
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        # 获取学生的日志
        query = WorkLog.query.filter_by(student_id=student_id)
        
        # 日期筛选
        if start_date:
            query = query.filter(WorkLog.date >= start_date)
        if end_date:
            query = query.filter(WorkLog.date <= end_date)
        
        logs = query.order_by(WorkLog.date.desc()).all()
        
        # 构建响应数据
        log_list = []
        for log in logs:
            log_list.append({
                'id': log.id,
                'content': log.content,
                'date': log.date.strftime('%Y-%m-%d'),
                'file': log.file,
                'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({'success': True, 'logs': log_list})
    except Exception as e:
        print(f"获取学生日志失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 查看日志附件
@app.route('/api/student/logs/file/<filename>', methods=['GET'])
@login_required
def view_log_file(filename):
    """查看日志附件"""
    try:
        # 安全处理文件名，防止路径遍历攻击
        import posixpath
        import ntpath
        # 规范化文件名，移除路径部分
        filename = posixpath.basename(filename)
        filename = ntpath.basename(filename)
        
        # 构建文件路径
        filepath = os.path.join(UPLOAD_FOLDER_LOGS, filename)
        print(f"尝试读取文件: {filepath}")
        
        if not os.path.exists(filepath):
            print(f"文件不存在: {filepath}")
            # 列出目录中的文件，方便调试
            files = os.listdir(UPLOAD_FOLDER_LOGS)
            print(f"目录中的文件: {files}")
            return jsonify({'success': False, 'message': '文件不存在'}), 404
        
        return send_file(filepath)
    except Exception as e:
        print(f"查看日志附件失败: {str(e)}")
        return jsonify({'success': False, 'message': f'查看失败: {str(e)}'})

# ==================== 团队管理API ====================

# 搜索教师
@app.route('/api/teachers/search', methods=['GET'])
@login_required
def search_teachers():
    """搜索教师"""
    try:
        keyword = request.args.get('keyword', '')
        
        if not keyword:
            return jsonify({'success': True, 'teachers': []})
        
        # 搜索教师，匹配用户名、真实姓名或专业
        keyword = f'%{keyword}%'
        teachers = User.query.filter(
            User.role == 'teacher',
            db.or_(
                User.username.like(keyword),
                User.realname.like(keyword),
                User.major.like(keyword)
            )
        ).all()
        
        # 构建响应数据
        teacher_list = []
        for teacher in teachers:
            teacher_list.append({
                'id': teacher.id,
                'username': teacher.username,
                'realname': teacher.realname,
                'major': teacher.major,
                'department': teacher.department
            })
        
        return jsonify({'success': True, 'teachers': teacher_list})
    except Exception as e:
        print(f"搜索教师失败: {str(e)}")
        return jsonify({'success': False, 'message': f'搜索失败: {str(e)}'})

# 创建团队申请
@app.route('/api/team/applications', methods=['POST'])
@login_required
def create_team_application():
    """创建团队申请"""
    try:
        user_id = request.current_user['user_id']
        data = request.get_json()
        
        theme = data.get('theme')
        courses = data.get('courses', [])
        members = data.get('members', [])
        description = data.get('description', '')
        
        if not theme or not courses or not members:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400
        
        # 检查所有课程是否都已分配教师
        for course in courses:
            if not course.get('teacherIds') or len(course['teacherIds']) == 0:
                return jsonify({'success': False, 'message': '所有课程都需要分配授课教师'}), 400
        
        # 将课程信息转换为字符串格式存储
        courses_str = ';'.join([f"{course['courseName']}:{','.join(map(str, course['teacherIds']))}" for course in courses])
        
        # 创建团队
        team = Team(
            theme=theme,
            course=courses_str,
            description=description,
            initiator_id=user_id,
            status='pending'
        )
        db.session.add(team)
        db.session.flush()  # 获取team.id
        
        # 添加团队成员（包括发起者）
        # 添加发起者
        initiator_member = TeamMember(
            team_id=team.id,
            user_id=user_id,
            status='accepted'  # 发起者自动接受
        )
        db.session.add(initiator_member)
        
        # 添加其他成员
        for member in members:
            member_id = member.get('id')
            if member_id and member_id != user_id:
                team_member = TeamMember(
                    team_id=team.id,
                    user_id=member_id,
                    status='pending'
                )
                db.session.add(team_member)
        
        db.session.commit()
        
        # 添加申请日志
        log = ApplicationLog(
            user_id=user_id,
            type='team',
            status='pending',
            receipt='团队组建申请',
            description=f'申请组建团队：{theme}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '团队申请创建成功'})
    except Exception as e:
        db.session.rollback()
        print(f"创建团队申请失败: {str(e)}")
        return jsonify({'success': False, 'message': f'创建失败: {str(e)}'})

# 获取我的团队
@app.route('/api/team/my-teams', methods=['GET'])
@login_required
def get_my_teams():
    """获取我的团队（支持分页和搜索）"""
    try:
        user_id = request.current_user['user_id']
        
        # 获取分页参数
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 10))
        keyword = request.args.get('keyword', '')
        
        # 获取用户参与的所有团队
        team_memberships = TeamMember.query.filter_by(user_id=user_id).all()
        team_ids = [tm.team_id for tm in team_memberships]
        
        # 获取这些团队的详细信息
        query = Team.query.filter(Team.id.in_(team_ids))
        
        # 关键词搜索
        if keyword:
            query = query.filter(Team.theme.like(f'%{keyword}%'))
        
        # 排序
        query = query.order_by(Team.create_time.desc())
        
        # 分页
        total_count = query.count()
        total_pages = (total_count + page_size - 1) // page_size
        teams = query.offset((page - 1) * page_size).limit(page_size).all()
        
        # 构建响应数据
        team_list = []
        for team in teams:
            # 获取团队成员信息
            member_info = []
            member_map = {}
            for member in team.members:
                user = User.query.get(member.user_id)
                if user:
                    member_info.append({
                        'id': user.id,
                        'realname': user.realname,
                        'username': user.username,
                        'status': member.status
                    })
                    member_map[user.id] = user.realname
            
            # 解析课程信息
            courses = []
            if team.course:
                for course_str in team.course.split(';'):
                    if ':' in course_str:
                        course_name, teacher_ids_str = course_str.split(':', 1)
                        teacher_ids = list(map(int, teacher_ids_str.split(',')))
                        teacher_names = [member_map.get(teacher_id, '') for teacher_id in teacher_ids]
                        courses.append({
                            'courseName': course_name,
                            'teacherIds': teacher_ids,
                            'teacherNames': teacher_names
                        })
            
            team_list.append({
                'id': team.id,
                'theme': team.theme,
                'courses': courses,
                'description': team.description,
                'status': team.status,
                'members': member_info,
                'create_time': team.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                'is_initiator': team.initiator_id == user_id
            })
        
        return jsonify({'success': True, 'teams': team_list, 'total_pages': total_pages})
    except Exception as e:
        print(f"获取我的团队失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 申请解散团队API
@app.route('/api/team/apply-dissolve', methods=['POST'])
@login_required
def apply_dissolve_team():
    """申请解散团队"""
    try:
        data = request.get_json()
        team_id = data.get('team_id')
        user_id = request.current_user['user_id']
        
        if not team_id:
            return jsonify({'success': False, 'message': '缺少团队ID'}), 400
        
        # 查找团队
        team = Team.query.get(team_id)
        if not team:
            return jsonify({'success': False, 'message': '团队不存在'}), 404
        
        # 检查是否为发起者
        if team.initiator_id != user_id:
            return jsonify({'success': False, 'message': '只有团队发起者才能申请解散'}), 403
        
        # 检查团队状态
        if team.status != 'approved':
            return jsonify({'success': False, 'message': '只能解散已批准的团队'}), 400
        
        # 更新团队状态为解散申请中
        team.status = 'dissolving'
        team.dissolve_time = datetime.utcnow()
        
        db.session.commit()
        
        # 添加申请日志
        log = ApplicationLog(
            user_id=user_id,
            type='team',
            status='pending',
            receipt='申请解散团队',
            description=f'申请解散团队：{team.theme}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '申请已提交，等待管理员审批'})
    except Exception as e:
        db.session.rollback()
        print(f"申请解散团队失败: {str(e)}")
        return jsonify({'success': False, 'message': f'申请失败: {str(e)}'})

# 获取对我的组队申请
@app.route('/api/team/my-applications', methods=['GET'])
@login_required
def get_my_team_applications():
    """获取对我的组队申请"""
    try:
        user_id = request.current_user['user_id']
        
        # 获取用户收到的组队申请（只显示team状态为pending的团队，但保留所有用户选择状态以允许修改）
        team_memberships = TeamMember.query.filter(
            TeamMember.user_id == user_id
        ).join(Team, TeamMember.team_id == Team.id).filter(
            Team.status == 'pending'
        ).all()
        
        # 构建响应数据
        application_list = []
        for tm in team_memberships:
            team = tm.team
            if team:
                # 获取发起者信息
                initiator = User.query.get(team.initiator_id)
                
                # 获取团队成员信息
                member_info = []
                member_map = {}
                for member in team.members:
                    user = User.query.get(member.user_id)
                    if user:
                        member_info.append({
                            'id': user.id,
                            'realname': user.realname
                        })
                        member_map[user.id] = user.realname
                
                # 解析课程信息
                courses = []
                if team.course:
                    for course_str in team.course.split(';'):
                        if ':' in course_str:
                            course_name, teacher_ids_str = course_str.split(':', 1)
                            teacher_ids = list(map(int, teacher_ids_str.split(',')))
                            teacher_names = [member_map.get(teacher_id, '') for teacher_id in teacher_ids]
                            courses.append({
                                'courseName': course_name,
                                'teacherIds': teacher_ids,
                                'teacherNames': teacher_names
                            })
                
                # 获取当前用户在该团队中的状态
                my_status = tm.status
                
                application_list.append({
                    'id': team.id,
                    'myStatus': my_status,
                    'theme': team.theme,
                    'courses': courses,
                    'description': team.description,
                    'initiator': {
                        'id': initiator.id,
                        'realname': initiator.realname
                    } if initiator else None,
                    'members': member_info,
                    'create_time': team.create_time.strftime('%Y-%m-%d %H:%M:%S')
                })
        
        return jsonify({'success': True, 'applications': application_list})
    except Exception as e:
        print(f"获取组队申请失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 处理组队申请
@app.route('/api/team/application/process', methods=['POST'])
@login_required
def process_team_application():
    """处理组队申请"""
    try:
        user_id = request.current_user['user_id']
        data = request.get_json()
        
        team_id = data.get('team_id')
        action = data.get('action')  # accept/reject
        
        if not team_id or action not in ['accept', 'reject']:
            return jsonify({'success': False, 'message': '参数错误'}), 400
        
        # 查找团队成员记录
        team_member = TeamMember.query.filter_by(
            team_id=team_id,
            user_id=user_id
        ).first()
        
        if not team_member:
            return jsonify({'success': False, 'message': '申请不存在'}), 404
        
        # 只有当团队状态为 'rejected' 或 'approved' 时，才不允许修改
        if team_member.team.status in ['rejected', 'approved']:
            return jsonify({'success': False, 'message': '申请已处理，无法修改'}), 400
        
        # 更新状态
        team_member.status = 'accepted' if action == 'accept' else 'rejected'
        
        # 检查是否所有成员都已接受
        team = team_member.team
        all_accepted = all(m.status == 'accepted' for m in team.members)
        any_rejected = any(m.status == 'rejected' for m in team.members)
        
        if all_accepted:
            team.status = 'pending'  # 所有成员接受后，等待管理员审批
        elif any_rejected:
            team.status = 'pending'  # 即使有成员拒绝，也等待管理员审批
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'申请已{"接受" if action == "accept" else "拒绝"}'})
    except Exception as e:
        db.session.rollback()
        print(f"处理组队申请失败: {str(e)}")
        return jsonify({'success': False, 'message': f'处理失败: {str(e)}'})

# 获取待审批的团队（管理员）
@app.route('/api/team/pending', methods=['GET'])
@login_required
def get_pending_teams():
    """获取待审批的团队"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        
        if user.role != 'admin':
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        # 获取待审批的团队（包括解散申请）
        teams = Team.query.filter(Team.status.in_(['pending', 'dissolving'])).all()
        
        # 构建响应数据
        team_list = []
        for team in teams:
            # 获取团队成员信息
            member_info = []
            member_map = {}
            for member in team.members:
                user = User.query.get(member.user_id)
                if user:
                    member_info.append({
                        'id': user.id,
                        'realname': user.realname,
                        'username': user.username,
                        'status': member.status
                    })
                    member_map[user.id] = user.realname
            
            # 解析课程信息
            courses = []
            if team.course:
                for course_str in team.course.split(';'):
                    if ':' in course_str:
                        course_name, teacher_ids_str = course_str.split(':', 1)
                        teacher_ids = list(map(int, teacher_ids_str.split(',')))
                        teacher_names = [member_map.get(teacher_id, '') for teacher_id in teacher_ids]
                        courses.append({
                            'courseName': course_name,
                            'teacherIds': teacher_ids,
                            'teacherNames': teacher_names
                        })
            
            team_list.append({
                'id': team.id,
                'theme': team.theme,
                'courses': courses,
                'description': team.description,
                'members': member_info,
                'create_time': team.create_time.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({'success': True, 'teams': team_list})
    except Exception as e:
        print(f"获取待审批团队失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 审批团队（管理员）
@app.route('/api/team/approve', methods=['POST'])
@login_required
def approve_team():
    """审批团队"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        
        if user.role != 'admin':
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        data = request.get_json()
        team_id = data.get('team_id')
        comment = data.get('comment', '')
        
        if not team_id:
            return jsonify({'success': False, 'message': '缺少团队ID'}), 400
        
        # 查找团队
        team = Team.query.get(team_id)
        if not team:
            return jsonify({'success': False, 'message': '团队不存在'}), 404
        
        # 根据当前状态决定批准类型
        if team.status == 'pending':
            # 批准组建团队
            team.status = 'approved'
            message = '团队已批准'
        elif team.status == 'dissolving':
            # 批准解散团队
            team.status = 'dissolved'
            message = '团队已解散'
        else:
            return jsonify({'success': False, 'message': '无效的团队状态'}), 400
        
        team.review_time = datetime.utcnow()
        team.review_comment = comment
        
        # 更新申请日志状态
        log = ApplicationLog.query.filter_by(
            user_id=team.initiator_id,
            type='team',
            status='pending',
            description=f'申请组建团队：{team.theme}'
        ).first()
        if not log:
            # 如果没有找到组建申请日志，尝试查找解散申请日志
            log = ApplicationLog.query.filter_by(
                user_id=team.initiator_id,
                type='team',
                status='pending',
                description=f'申请解散团队：{team.theme}'
            ).first()
        if log:
            log.status = 'approved'
            log.handler_id = user.id
            log.handler_name = user.realname
            log.receipt = message
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        db.session.rollback()
        print(f"批准团队失败: {str(e)}")
        return jsonify({'success': False, 'message': f'批准失败: {str(e)}'})

# 拒绝团队（管理员）
@app.route('/api/team/reject', methods=['POST'])
@login_required
def reject_team():
    """拒绝团队或驳回解散申请"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        
        if user.role != 'admin':
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        data = request.get_json()
        team_id = data.get('team_id')
        comment = data.get('comment', '')
        
        if not team_id:
            return jsonify({'success': False, 'message': '缺少团队ID'}), 400
        
        # 查找团队
        team = Team.query.get(team_id)
        if not team:
            return jsonify({'success': False, 'message': '团队不存在'}), 404
        
        # 根据当前状态决定拒绝类型
        if team.status == 'pending':
            # 拒绝组建团队
            team.status = 'rejected'
            message = '团队已拒绝'
        elif team.status == 'dissolving':
            # 驳回解散申请
            team.status = 'approved'
            team.dissolve_reason = None
            team.dissolve_time = None
            message = '解散申请已驳回'
        else:
            return jsonify({'success': False, 'message': '无效的团队状态'}), 400
        
        team.review_time = datetime.utcnow()
        team.review_comment = comment
        
        # 更新申请日志状态
        log = ApplicationLog.query.filter_by(
            user_id=team.initiator_id,
            type='team',
            status='pending',
            description=f'申请组建团队：{team.theme}'
        ).first()
        if not log:
            # 如果没有找到组建申请日志，尝试查找解散申请日志
            log = ApplicationLog.query.filter_by(
                user_id=team.initiator_id,
                type='team',
                status='pending',
                description=f'申请解散团队：{team.theme}'
            ).first()
        if log:
            log.status = 'rejected'
            log.handler_id = user.id
            log.handler_name = user.realname
            log.receipt = message
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        db.session.rollback()
        print(f"拒绝团队失败: {str(e)}")
        return jsonify({'success': False, 'message': f'拒绝失败: {str(e)}'})

# 编辑日志
@app.route('/api/student/logs/<int:log_id>', methods=['PUT'])
@login_required
def update_student_log(log_id):
    """编辑学生日志"""
    try:
        user_id = request.current_user['user_id']
        
        # 查找日志
        log = WorkLog.query.filter_by(id=log_id, student_id=user_id).first()
        if not log:
            return jsonify({'success': False, 'message': '日志不存在或无权限编辑'}), 404
        
        # 检查是否是最新的日志
        latest_log = WorkLog.query.filter_by(student_id=user_id).order_by(WorkLog.date.desc(), WorkLog.created_at.desc()).first()
        if latest_log.id != log_id:
            return jsonify({'success': False, 'message': '只能编辑最新发布的日志'}), 403
        
        # 检查请求是否包含文件
        file = request.files.get('file')
        
        content = request.form.get('content')
        date = request.form.get('date')
        
        if not content or not date:
            return jsonify({'success': False, 'message': '缺少必填字段'}), 400
        
        # 保存文件（如果有）
        if file and file.filename:
            import uuid
            # 清理文件名，移除特殊字符
            import re
            clean_filename = re.sub(r'[<>:"/\\|?*]', '_', file.filename)
            filename = f"{uuid.uuid4()}_{clean_filename}"
            filepath = os.path.join(UPLOAD_FOLDER_LOGS, filename)
            file.save(filepath)
            
            # 删除旧文件（如果有）
            if log.file:
                old_filepath = os.path.join(UPLOAD_FOLDER_LOGS, log.file)
                if os.path.exists(old_filepath):
                    os.remove(old_filepath)
            
            log.file = filename
        
        # 更新日志内容
        log.content = content
        log.date = date
        log.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': '日志更新成功'})
    except Exception as e:
        db.session.rollback()
        # 清理上传的文件
        if 'filepath' in locals() and os.path.exists(filepath):
            os.remove(filepath)
        print(f"更新日志失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'})






# 校内用户注册API（管理员使用）
@app.route('/api/innerregister', methods=['POST'])
@login_required
def inner_register():
    # 检查是否为管理员
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以注册校内用户'}), 403
    
    data = request.get_json()
    username = data.get('username')
    realname = data.get('realname')
    role = data.get('role', 'student')  # 默认为学生
    id_card = data.get('id_card')  # 现在接收完整身份证号
    major = data.get('major')
    department = data.get('department')
    contact = data.get('contact')
    title = data.get('title')
    email = data.get('email')
    
    if not username or not realname or not id_card:
        return jsonify({'success': False, 'message': '缺少必要字段（用户名、真实姓名和身份证号）'}), 400
    
    # 自动截取身份证后六位
    id_card_last6 = str(id_card)[-6:]
    
    # 检查用户名是否已存在
    if User.query.filter_by(username=username).first():
        return jsonify({'success': False, 'message': '用户名已存在'}), 400
    
    # 使用身份证后六位作为默认密码，首次登录时会强制修改
    password_hash = bcrypt.generate_password_hash(id_card_last6).decode('utf-8')
    
    # 自动生成邮箱
    if role == 'student':
        email = f"{username}@neu.edu.cn"
    elif role == 'teacher':
        # 教师邮箱使用真实姓名
        # 移除姓名中的空格，确保邮箱格式正确
        clean_realname = realname.replace(' ', '')
        email = f"{clean_realname}@neu.edu.cn"
    else:
        email = email  # 管理员保持原有邮箱
    
    # 创建用户
    user = User(
        username=username,
        password_hash=password_hash,
        role=role,
        realname=realname,
        id_card_last6=id_card_last6,
        major=major,
        department=department,
        contact=contact,
        title=title,
        email=email,
        first_login=True,
        is_confirmed=True  # 校内注册直接确认
    )
    
    try:
        db.session.add(user)
        db.session.commit()
        return jsonify({'success': True, 'message': '校内用户注册成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'注册失败: {str(e)}'}), 500




# 批量注册API（管理员使用）
@app.route('/api/download-template', methods=['GET'])
@login_required
def download_template():
    """下载批量注册模板"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以下载模板'}), 403
    
    try:
        # 导入openpyxl库
        import openpyxl
        import io
        
        # 创建一个新的Excel工作簿
        workbook = openpyxl.Workbook()
        
        # 创建学生模板sheet
        student_sheet = workbook.active
        student_sheet.title = '学生模板'
        
        # 学生表头
        student_headers = ['username', 'realname', 'role', 'id_card', 'department', 'major', 'contact']
        student_chinese_headers = ['学号/工号', '真实姓名', '角色（student）', '身份证号（18位）', '院系', '专业', '联系方式']
        
        # 写入学生表头
        for i, (header, chinese_header) in enumerate(zip(student_headers, student_chinese_headers), 1):
            student_sheet.cell(row=1, column=i, value=f"{header} ({chinese_header})")
        
        # 写入学生示例数据
        student_sample_data = [
            ['20230001', '张三', 'student', '110101199001011234', '计算机学院', '计算机科学与技术', '13800138001'],
            ['20230002', '李四', 'student', '110101199001015678', '电子学院', '电子信息工程', '13900139002']
        ]
        
        for row_idx, row_data in enumerate(student_sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                student_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 创建教师模板sheet
        teacher_sheet = workbook.create_sheet('教师模板')
        
        # 教师表头
        teacher_headers = ['username', 'realname', 'role', 'id_card', 'department', 'major', 'contact', 'title']
        teacher_chinese_headers = ['工号', '真实姓名', '角色（teacher）', '身份证号（18位）', '院系', '专业', '联系方式', '职称']
        
        # 写入教师表头
        for i, (header, chinese_header) in enumerate(zip(teacher_headers, teacher_chinese_headers), 1):
            teacher_sheet.cell(row=1, column=i, value=f"{header} ({chinese_header})")
        
        # 写入教师示例数据
        teacher_sample_data = [
            ['T2023001', '王五', 'teacher', '110101198001011234', '计算机学院', '计算机科学与技术', '13800138003', '教授'],
            ['T2023002', '赵六', 'teacher', '110101198001015678', '电子学院', '电子信息工程', '13900139004', '副教授']
        ]
        
        for row_idx, row_data in enumerate(teacher_sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                teacher_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # 创建临时文件
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # 返回文件
        return send_file(output, as_attachment=True, download_name='批量注册模板.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'success': False, 'message': f'生成模板失败: {str(e)}'}), 500

@app.route('/api/batch-register', methods=['POST'])
@login_required
def batch_register():
    # 检查是否为管理员
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以批量注册用户'}), 403
    
    # 检查是否有文件上传
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '缺少文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '文件名为空'}), 400
    
    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '只支持Excel文件'}), 400
    
    try:
        # 导入openpyxl库
        import openpyxl
        
        # 加载Excel文件
        workbook = openpyxl.load_workbook(file)
        
        # 中文表头映射
        header_mapping = {
            '学号/工号': 'username',
            '真实姓名': 'realname', 
            '角色': 'role',
            '角色（student/teacher/admin）': 'role',
            '身份证号': 'id_card',
            '身份证号（18位）': 'id_card',
            '院系': 'department',
            '专业': 'major',
            '联系方式': 'contact',
            '职称': 'title'
        }
        
        # 批量注册用户
        registered_count = 0
        errors = []
        
        # 遍历所有sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 读取表头
            headers = [cell.value for cell in sheet[1]]
            
            # 转换表头
            mapped_headers = []
            for header in headers:
                if header in header_mapping:
                    mapped_headers.append(header_mapping[header])
                else:
                    # 尝试从"username (学号/工号)"格式中提取字段名
                    if ' (' in header and ')' in header:
                        # 提取括号前的部分作为字段名
                        field_name = header.split(' (')[0].strip()
                        mapped_headers.append(field_name)
                    else:
                        # 直接使用原表头
                        mapped_headers.append(header)
            
            required_headers = ['username', 'realname', 'role']
            for header in required_headers:
                if header not in mapped_headers:
                    errors.append(f'Sheet "{sheet_name}": 缺少必要列: {header}')
                    continue
            
            # 处理当前sheet中的数据
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(mapped_headers, row))
                
                username = row_data.get('username')
                realname = row_data.get('realname')
                role = row_data.get('role', 'student')
                id_card = row_data.get('id_card')  # 现在接收完整身份证号
                major = row_data.get('major')
                department = row_data.get('department')
                contact = row_data.get('contact')
                title = row_data.get('title')
                email = row_data.get('email')
                
                if not username or not realname:
                    errors.append(f'Sheet "{sheet_name}" 第{sheet._current_row-1}行: 缺少用户名或真实姓名')
                    continue
                
                if not id_card:
                    errors.append(f'Sheet "{sheet_name}" 第{sheet._current_row-1}行: 缺少身份证号')
                    continue
                
                # 自动截取身份证后六位
                id_card_last6 = str(id_card)[-6:]
                
                # 检查用户名是否已存在
                if User.query.filter_by(username=username).first():
                    errors.append(f'Sheet "{sheet_name}" 第{sheet._current_row-1}行: 用户名 {username} 已存在')
                    continue
                
                # 设置默认密码为身份证后六位，首次登录时会强制修改
                password_hash = bcrypt.generate_password_hash(str(id_card_last6)).decode('utf-8')
                
                # 自动生成邮箱
                if role == 'student':
                    email = f"{username}@neu.edu.cn"
                elif role == 'teacher':
                    # 教师邮箱使用真实姓名
                    # 移除姓名中的空格，确保邮箱格式正确
                    clean_realname = str(realname).replace(' ', '')
                    email = f"{clean_realname}@neu.edu.cn"
                else:
                    email = email  # 管理员保持原有邮箱
                
                # 创建用户
                user = User(
                    username=username,
                    password_hash=password_hash,
                    role=role,
                    realname=realname,
                    id_card_last6=id_card_last6,
                    major=major,
                    department=department,
                    contact=contact,
                    title=title,
                    email=email,
                    first_login=True,
                    is_confirmed=True  # 校内注册直接确认
                )
                
                db.session.add(user)
                registered_count += 1
        
        # 提交事务
        db.session.commit()
        
        # 统计重复用户数量
        duplicate_count = 0
        other_errors_count = 0
        
        for error in errors:
            if '已存在' in error:
                duplicate_count += 1
            else:
                other_errors_count += 1
        
        if errors:
            # 只要有错误，就标记为注册失败
            message = f'批量注册失败，成功注册 {registered_count} 个用户'
            if duplicate_count > 0:
                message += f'，重复 {duplicate_count} 个用户'
            if other_errors_count > 0:
                message += f'，其他错误 {other_errors_count} 个'
            
            return jsonify({
                'success': False,
                'message': message,
                'registered_count': registered_count,
                'duplicate_count': duplicate_count,
                'other_errors_count': other_errors_count,
                'errors': errors
            })
        else:
            return jsonify({
                'success': True,
                'message': f'批量注册成功，共注册 {registered_count} 个用户',
                'registered_count': registered_count
            })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'批量注册失败: {str(e)}'}), 500

# 登录API
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    student_id = data.get('studentId')  # 学号/工号
    id_card_last6 = data.get('idCardLast6')  # 身份证后六位
    
    if not student_id or not id_card_last6:
        return jsonify({'success': False, 'message': '学号/工号和身份证后六位不能为空'}), 400
    
    
    
    # 查找用户
    user = User.query.filter_by(username=student_id).first()
    
    if not user:
        return jsonify({'success': False, 'message': '学号/工号不存在'}), 401
    
    # 检查企业用户是否已激活
    if user.role == 'company' and not user.is_active:
        # 企业用户未激活，可以登录但提示需要审批
        token = generate_token(user.id, user.username, user.role)
        return jsonify({
            'success': True,
            'requireChangePassword': False,
            'message': '账号等待管理员审批，部分功能受限',
            'token': token,
            'user': {
                'id': user.id,
                'username': user.username,
                'realname': user.realname,
                'role': user.role,
                'is_active': False
            },
            'is_pending': True
        })
    
    # 检查是否首次登录
    if user.first_login:
        # 首次登录：验证身份证后六位
        if user.id_card_last6 != id_card_last6:
            return jsonify({'success': False, 'message': '身份证后六位错误'}), 401
        
        # 首次登录成功，生成 JWT Token
        token = generate_token(user.id, user.username, user.role)
        
        return jsonify({
            'success': True,
            'requireChangePassword': True,
            'message': '首次登录，请修改密码',
            'token': token,
            'user': {
                'id': user.id,
                'username': user.username,
                'realname': user.realname,
                'role': user.role,
                'major': user.major,
                'is_active': user.is_active
            }
        })
    else:
        # 非首次登录：验证密码
        if not bcrypt.check_password_hash(user.password_hash, id_card_last6):
            return jsonify({'success': False, 'message': '密码错误'}), 401
        
        # 登录成功，生成 JWT Token
        token = generate_token(user.id, user.username, user.role)
        
        return jsonify({
            'success': True,
            'requireChangePassword': False,
            'message': '登录成功',
            'token': token,
            'user': {
                'id': user.id,
                'username': user.username,
                'realname': user.realname,
                'role': user.role,
                'major': user.major,
                'is_active': user.is_active
            }
        })

# 获取待审批的企业用户列表（管理员使用）
@app.route('/api/companies/pending', methods=['GET'])
@login_required
def get_pending_companies():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以查看待审批企业'}), 403
    
    pending_companies = Company.query.filter_by(status='pending').all()
    
    result = []
    for company in pending_companies:
        result.append({
            'id': company.id,
            'user_id': company.user_id,
            'username': company.user.username,
            'realname': company.user.realname,
            'company_name': company.company_name,
            'field': company.field,
            'nature': company.nature,
            'scale': company.scale,
            'contact': company.contact,
            'proof_file': company.proof_file,
            'create_time': company.create_time.strftime('%Y-%m-%d %H:%M:%S') if company.create_time else None
        })
    
    return jsonify({'success': True, 'companies': result})

# 审批企业用户（管理员使用）
@app.route('/api/companies/approve', methods=['POST'])
@login_required
def approve_company():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以审批企业'}), 403
    
    data = request.get_json()
    company_id = data.get('company_id')
    action = data.get('action')  # approve/reject
    comment = data.get('comment', '')
    
    if not company_id or action not in ['approve', 'reject']:
        return jsonify({'success': False, 'message': '参数错误'}), 400
    
    company = Company.query.get(company_id)
    if not company:
        return jsonify({'success': False, 'message': '企业不存在'}), 404
    
    if company.status != 'pending':
        return jsonify({'success': False, 'message': '该企业已经被审批过'}), 400
    
    try:
        company.status = 'approved' if action == 'approve' else 'rejected'
        company.review_time = datetime.utcnow()
        company.review_comment = comment
        
        if action == 'approve':
            # 激活用户账号
            company.user.is_active = True
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'企业已{"批准" if action == "approve" else "拒绝"}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'操作失败: {str(e)}'}), 500

# 下载企业证明材料
@app.route('/api/companies/proof/<filename>', methods=['GET'])
@login_required
def download_company_proof(filename):
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以下载证明材料'}), 403
    
    filepath = os.path.join(UPLOAD_FOLDER_COMPANY, filename)
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'message': '文件不存在'}), 404
    
    # 编码文件名，解决中文问题
    import urllib.parse
    encoded_filename = urllib.parse.quote(filename)
    
    return send_file(filepath, as_attachment=True, download_name=encoded_filename)

#更改联系方式api
@app.route('/api/user/profile', methods=['PUT'])
@login_required
def update_user_profile():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    try:
        data = request.get_json()
        contact = data.get('contact')
        email = data.get('email')
        
        # 验证邮箱格式
        if email:
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, email):
                return jsonify({'success': False, 'message': '邮箱格式不正确'}), 400
        
        # 更新联系方式
        if contact is not None:
            user.contact = contact
        if email is not None:
            user.email = email
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '联系方式更新成功',
            'user': {
                'contact': user.contact,
                'email': user.email
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'}), 500

# 获取用户申请日志API
@app.route('/api/user/application-logs', methods=['GET'])
@login_required
def get_user_application_logs():
    """获取用户的申请日志（支持分页和筛选）"""
    try:
        user_id = request.current_user['user_id']
        
        # 获取分页参数
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 10))
        
        # 获取筛选参数
        keyword = request.args.get('keyword', '')
        types = request.args.get('types', '').split(',') if request.args.get('types') else []
        statuses = request.args.get('statuses', '').split(',') if request.args.get('statuses') else []
        
        # 构建查询
        query = ApplicationLog.query.filter_by(user_id=user_id)
        
        # 类型筛选
        if types and types != ['']:
            query = query.filter(ApplicationLog.type.in_(types))
        
        # 状态筛选
        if statuses and statuses != ['']:
            query = query.filter(ApplicationLog.status.in_(statuses))
        
        # 关键词搜索（搜索描述和回执信息）
        if keyword:
            query = query.filter(
                (ApplicationLog.description.like(f'%{keyword}%')) |
                (ApplicationLog.receipt.like(f'%{keyword}%'))
            )
        
        # 排序
        query = query.order_by(ApplicationLog.create_time.desc())
        
        # 分页
        total_count = query.count()
        total_pages = (total_count + page_size - 1) // page_size
        logs = query.offset((page - 1) * page_size).limit(page_size).all()
        
        # 构建响应数据
        log_list = []
        for log in logs:
            log_list.append({
                'id': log.id,
                'type': log.type,
                'status': log.status,
                'handler_name': log.handler_name,
                'receipt': log.receipt,
                'description': log.description,
                'create_time': log.create_time.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({'success': True, 'logs': log_list, 'total_pages': total_pages, 'total_count': total_count})
    except Exception as e:
        print(f"获取申请日志失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 修改密码API
@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.get_json()
    old_password = data.get('old_password')
    new_password = data.get('new_password')
    
    if not new_password:
        return jsonify({'success': False, 'message': '新密码不能为空'}), 400
    
    if len(new_password) < 6:
        return jsonify({'success': False, 'message': '密码至少需要6位'}), 400
    
    # 直接从请求头获取token并验证
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'success': False, 'message': '缺少认证令牌'}), 401
    
    token = auth_header.split(' ')[1]
    payload = decode_token(token)
    if not payload:
        return jsonify({'success': False, 'message': '令牌无效或已过期'}), 401
    
    user_id = payload.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': '令牌格式无效'}), 401
    
    user = User.query.get(user_id)
    
    # 如果提供了旧密码，验证旧密码
    if old_password:
        if not bcrypt.check_password_hash(user.password_hash, old_password):
            return jsonify({'success': False, 'message': '旧密码不正确'}), 400
    
    # 加密新密码
    password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
    
    try:
        user.password_hash = password_hash
        user.first_login = False  # 标记为非首次登录
        db.session.commit()
        
        # 清除修改密码标记
        session.pop('require_change_password', None)
        
        return jsonify({
            'success': True,
            'message': '密码修改成功'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'修改失败: {str(e)}'}), 500

# 登出API
@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True, 'message': '登出成功'})

# 检查登录状态API (JWT 版本)
@app.route('/api/check-auth', methods=['GET'])
def check_auth():
    token = get_token_from_header()
    
    if not token:
        return jsonify({'success': False, 'message': '未登录'}), 401
    
    payload = decode_token(token)
    if not payload:
        return jsonify({'success': False, 'message': '令牌无效或已过期'}), 401
    
    user = User.query.get(payload['user_id'])
    if user:
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'realname': user.realname,
                'role': user.role,
                'major': user.major,
                'contact': user.contact,
                'email': user.email
            }
        })
    
    return jsonify({'success': False, 'message': '用户不存在'}), 404

# 获取用户的申请记录（pending状态）
@app.route('/api/my-applications', methods=['GET'])
@login_required
def get_my_applications():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    
    try:
        applications = []
        
        # 获取科研训练申请（pending状态）
        if user.role == 'student':
            research_applications = ResearchRegistration.query.filter_by(
                student_id=user_id,
                status='pending'
            ).all()
            
            for app in research_applications:
                project = ResearchProject.query.get(app.project_id)
                if project:
                    applications.append({
                        'id': app.id,
                        'type': 'research',
                        'name': project.project_name,
                        'status': app.status
                    })
        
        # 获取双选申请（pending状态）
        if user.role == 'student':
            # 学生的双选申请
            selection_applications = TeacherSelection.query.filter_by(
                student_id=user_id,
                status='pending'
            ).all()
            
            for app in selection_applications:
                teacher = User.query.get(app.teacher_id)
                if teacher:
                    applications.append({
                        'id': app.id,
                        'type': 'selection',
                        'name': teacher.realname,
                        'status': app.status
                    })
        elif user.role == 'teacher':
            # 教师收到的双选申请
            selection_applications = TeacherSelection.query.filter_by(
                teacher_id=user_id,
                status='pending'
            ).all()
            
            for app in selection_applications:
                student = User.query.get(app.student_id)
                if student:
                    applications.append({
                        'id': app.id,
                        'type': 'selection',
                        'name': student.realname,
                        'status': app.status
                    })
        
        return jsonify({
            'success': True,
            'applications': applications
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取申请记录失败: {str(e)}'}), 500

# ==================== 导师双选API ====================

# 获取当前双选阶段
@app.route('/api/double-selection/step', methods=['GET'])
@login_required
def get_double_selection_step():
    step = DoubleSelectionStep.query.first()
    if not step:
        step = DoubleSelectionStep(current_step=0)
        db.session.add(step)
        db.session.commit()
    
    return jsonify({
        'success': True,
        'step': step.current_step
    })

# 获取双选阶段时间
@app.route('/api/double-selection/time', methods=['GET'])
@login_required
def get_double_selection_time():
    # 获取所有阶段的时间信息
    time_records = DoubleSelectionTime.query.all()
    times = {}
    for record in time_records:
        times[record.step] = {
            'startTime': record.start_time.strftime('%Y-%m-%d %H:%M:%S') if record.start_time else None,
            'endTime': record.end_time.strftime('%Y-%m-%d %H:%M:%S') if record.end_time else None
        }
    
    return jsonify({
        'success': True,
        'times': times
    })

# 设置双选阶段（仅管理员）
@app.route('/api/double-selection/step', methods=['POST'])
@login_required
def set_double_selection_step():
    # 从请求上下文中获取用户信息（JWT方式）
    user_id = request.current_user['user_id']
    user = db.session.get(User, user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以设置双选阶段'}), 403
    
    data = request.get_json()
    new_step = data.get('step')
    start_time = data.get('startTime')
    end_time = data.get('endTime')
    
    if new_step is None or not isinstance(new_step, int) or new_step < 0 or new_step > 3:
        return jsonify({'success': False, 'message': '无效的阶段值，必须是0-3之间的整数'}), 400
    
    step = DoubleSelectionStep.query.first()
    if not step:
        step = DoubleSelectionStep(current_step=new_step)
        db.session.add(step)
    else:
        step.current_step = new_step
    
    # 设置阶段时间
    if start_time or end_time:
        time_record = DoubleSelectionTime.query.filter_by(step=new_step).first()
        if not time_record:
            time_record = DoubleSelectionTime(step=new_step)
            db.session.add(time_record)
        
        if start_time:
            time_record.start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
        if end_time:
            time_record.end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '双选阶段设置成功',
        'step': step.current_step
    })

# 获取导师列表
@app.route('/api/teachers', methods=['GET'])
@login_required
def get_teachers():
    """获取导师列表"""
    # 从双选导师表中获取参与双选的导师
    double_teachers = DoubleSelectionTeacher.query.filter_by(status='active').all()
    
    result = []
    for dt in double_teachers:
        teacher = dt.teacher
        result.append({
            'id': teacher.id,
            'username': teacher.username,
            'realname': teacher.realname,
            'major': teacher.major,
            'department': teacher.department,
            'title': teacher.title,
            'min_quota': dt.min_quota,
            'max_quota': dt.max_quota,
            'current_quota': dt.current_quota or 0
        })
    
    return jsonify({'success': True, 'teachers': result})

@app.route('/api/double-selection/teachers/join', methods=['POST'])
@login_required
def join_double_selection():
    """导师报名参与双选"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    # 检查是否为教师
    if user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以报名参与双选'}), 403
    
    # 检查是否已经报名
    existing = DoubleSelectionTeacher.query.filter_by(teacher_id=user_id).first()
    if existing:
        return jsonify({'success': False, 'message': '您已经报名参与双选'}), 400
    
    # 获取双选阶段
    step = DoubleSelectionStep.query.first()
    current_step = step.current_step if step else 0
    if current_step != 0:
        return jsonify({'success': False, 'message': '当前不在报名阶段'}), 400
    
    # 获取请求数据
    data = request.get_json()
    min_quota = data.get('min_quota', 0)
    max_quota = data.get('max_quota', 5)  # 默认最大名额为5
    
    # 创建双选导师记录
    double_teacher = DoubleSelectionTeacher(
        teacher_id=user_id,
        min_quota=min_quota,
        max_quota=max_quota,
        current_quota=0
    )
    
    try:
        db.session.add(double_teacher)
        db.session.commit()
        return jsonify({'success': True, 'message': '报名成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'报名失败: {str(e)}'}), 500

@app.route('/api/double-selection/teachers/add', methods=['POST'])
@login_required
def add_double_selection_teacher():
    """管理员指定导师参与双选"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    # 检查是否为管理员
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以指定导师'}), 403
    
    # 获取双选阶段
    step = DoubleSelectionStep.query.first()
    current_step = step.current_step if step else 0
    if current_step != 0:
        return jsonify({'success': False, 'message': '当前不在报名阶段'}), 400
    
    # 获取请求数据
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    min_quota = data.get('min_quota', 0)
    max_quota = data.get('max_quota', 5)  # 默认最大名额为5
    
    if not teacher_id:
        return jsonify({'success': False, 'message': '缺少教师ID'}), 400
    
    # 检查教师是否存在
    teacher = User.query.filter_by(id=teacher_id, role='teacher').first()
    if not teacher:
        return jsonify({'success': False, 'message': '教师不存在'}), 404
    
    # 检查是否已经报名
    existing = DoubleSelectionTeacher.query.filter_by(teacher_id=teacher_id).first()
    if existing:
        return jsonify({'success': False, 'message': '该教师已经参与双选'}), 400
    
    # 创建双选导师记录
    double_teacher = DoubleSelectionTeacher(
        teacher_id=teacher_id,
        min_quota=min_quota,
        max_quota=max_quota,
        current_quota=0
    )
    
    try:
        db.session.add(double_teacher)
        db.session.commit()
        return jsonify({'success': True, 'message': '添加成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'}), 500

@app.route('/api/double-selection/teachers/list', methods=['GET'])
@login_required
def get_double_selection_teachers():
    """获取参与双选的导师列表"""
    double_teachers = DoubleSelectionTeacher.query.filter_by(status='active').all()
    
    result = []
    for dt in double_teachers:
        teacher = dt.teacher
        result.append({
            'id': teacher.id,
            'username': teacher.username,
            'realname': teacher.realname,
            'major': teacher.major,
            'department': teacher.department,
            'title': teacher.title,
            'min_quota': dt.min_quota,
            'max_quota': dt.max_quota,
            'current_quota': dt.current_quota or 0,
            'join_time': dt.join_time.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return jsonify({'success': True, 'teachers': result})

@app.route('/api/double-selection/teachers/remove', methods=['POST'])
@login_required
def remove_double_selection_teacher():
    """管理员移除参与双选的导师"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    # 检查是否为管理员
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以移除导师'}), 403
    
    # 获取双选阶段
    step = DoubleSelectionStep.query.first()
    current_step = step.current_step if step else 0
    if current_step != 0:
        return jsonify({'success': False, 'message': '当前不在报名阶段'}), 400
    
    # 获取请求数据
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    
    if not teacher_id:
        return jsonify({'success': False, 'message': '缺少教师ID'}), 400
    
    # 查找双选导师记录
    double_teacher = DoubleSelectionTeacher.query.filter_by(teacher_id=teacher_id).first()
    if not double_teacher:
        return jsonify({'success': False, 'message': '该教师未参与双选'}), 404
    
    try:
        db.session.delete(double_teacher)
        db.session.commit()
        return jsonify({'success': True, 'message': '移除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'移除失败: {str(e)}'}), 500

@app.route('/api/double-selection/teachers/update', methods=['POST'])
@login_required
def update_double_selection_teacher():
    """管理员更新参与双选的导师名额"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    # 检查是否为管理员
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以更新导师名额'}), 403
    
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    min_quota = data.get('min_quota')
    max_quota = data.get('max_quota')
    
    if not teacher_id:
        return jsonify({'success': False, 'message': '缺少教师ID'}), 400
    
    try:
        # 查找双选导师记录
        double_teacher = DoubleSelectionTeacher.query.filter_by(teacher_id=teacher_id).first()
        if not double_teacher:
            return jsonify({'success': False, 'message': '导师未参与双选'}), 404
        
        # 更新名额
        if min_quota is not None:
            double_teacher.min_quota = min_quota
        if max_quota is not None:
            double_teacher.max_quota = max_quota
        
        db.session.commit()
        return jsonify({'success': True, 'message': '名额更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'更新失败: {str(e)}'}), 500

# 搜索用户接口
@app.route('/api/users', methods=['GET'])
@login_required
def search_users():
    """搜索用户，支持按角色和关键词搜索"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    # 检查是否为管理员
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以搜索用户'}), 403
    
    # 获取查询参数
    role = request.args.get('role')
    keyword = request.args.get('keyword', '')
    
    # 构建查询
    query = User.query
    
    # 按角色筛选
    if role:
        query = query.filter_by(role=role)
    
    # 按关键词筛选（匹配用户名、真实姓名或专业）
    if keyword:
        keyword = f'%{keyword}%'
        query = query.filter(
            db.or_(
                User.username.like(keyword),
                User.realname.like(keyword),
                User.major.like(keyword)
            )
        )
    
    # 执行查询
    users = query.all()
    
    # 构建返回数据
    users_list = []
    for user in users:
        users_list.append({
            'id': user.id,
            'username': user.username,
            'realname': user.realname,
            'major': user.major,
            'role': user.role
        })
    
    return jsonify({'success': True, 'users': users_list})

# 学生参与双选相关接口
@app.route('/api/double-selection/students/list', methods=['GET'])
@login_required
def get_double_selection_students():
    """获取参与双选的学生列表"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    # 检查是否为管理员
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以查看学生列表'}), 403
    
    # 获取所有参与双选的学生
    double_students = DoubleSelectionStudent.query.all()
    students = []
    
    for ds in double_students:
        student = User.query.get(ds.student_id)
        if student:
            students.append({
                'id': student.id,
                'username': student.username,
                'realname': student.realname,
                'major': student.major
            })
    
    return jsonify({'success': True, 'students': students})

@app.route('/api/double-selection/students/add', methods=['POST'])
@login_required
def add_double_selection_student():
    """管理员添加学生参与双选"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    # 检查是否为管理员
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以添加学生'}), 403
    
    # 获取双选阶段
    step = DoubleSelectionStep.query.first()
    current_step = step.current_step if step else 0
    if current_step != 0:
        return jsonify({'success': False, 'message': '当前不在报名阶段'}), 400
    
    # 获取请求数据
    data = request.get_json()
    student_id = data.get('student_id')
    
    if not student_id:
        return jsonify({'success': False, 'message': '缺少学生ID'}), 400
    
    # 检查学生是否存在
    student = User.query.get(student_id)
    if not student or student.role != 'student':
        return jsonify({'success': False, 'message': '学生不存在或不是学生角色'}), 404
    
    # 检查是否已经参与双选
    existing = DoubleSelectionStudent.query.filter_by(student_id=student_id).first()
    if existing:
        return jsonify({'success': False, 'message': '该学生已经参与双选'}), 400
    
    try:
        # 添加学生参与双选
        double_student = DoubleSelectionStudent(student_id=student_id)
        db.session.add(double_student)
        db.session.commit()
        return jsonify({'success': True, 'message': '学生添加成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'}), 500

@app.route('/api/double-selection/students/remove', methods=['POST'])
@login_required
def remove_double_selection_student():
    """管理员移除参与双选的学生"""
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    # 检查是否为管理员
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以移除学生'}), 403
    
    # 获取双选阶段
    step = DoubleSelectionStep.query.first()
    current_step = step.current_step if step else 0
    if current_step != 0:
        return jsonify({'success': False, 'message': '当前不在报名阶段'}), 400
    
    # 获取请求数据
    data = request.get_json()
    student_id = data.get('student_id')
    
    if not student_id:
        return jsonify({'success': False, 'message': '缺少学生ID'}), 400
    
    # 查找双选学生记录
    double_student = DoubleSelectionStudent.query.filter_by(student_id=student_id).first()
    if not double_student:
        return jsonify({'success': False, 'message': '该学生未参与双选'}), 404
    
    try:
        db.session.delete(double_student)
        db.session.commit()
        return jsonify({'success': True, 'message': '移除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'移除失败: {str(e)}'}), 500

# 获取我的选择
@app.route('/api/my-selection', methods=['GET'])
@login_required
def get_my_selection():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'student':
        return jsonify({'success': False, 'message': '只有学生可以查看选择'}), 403
    
    # 获取学生的所有选择（按志愿顺序）
    selections = TeacherSelection.query.filter_by(student_id=user_id).order_by(TeacherSelection.priority).all()
    
    if selections:
        selections_list = []
        for selection in selections:
            selections_list.append({
                'id': selection.id,
                'teacherId': selection.teacher_id,
                'teacherName': selection.teacher.realname,
                'priority': selection.priority,
                'status': selection.status,
                'createTime': selection.create_time.strftime('%Y-%m-%d %H:%M:%S')
            })
        return jsonify({
            'success': True,
            'selections': selections_list
        })
    
    return jsonify({
        'success': True,
        'selections': []
    })

# 学生选择导师
@app.route('/api/select-teacher', methods=['POST'])
@login_required
def select_teacher():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'student':
        return jsonify({'success': False, 'message': '只有学生可以选择导师'}), 403
    
    # 检查双选阶段
    step = DoubleSelectionStep.query.first()
    if not step or step.current_step != 1:
        return jsonify({'success': False, 'message': '当前不在选导师阶段'}), 400
    
    data = request.get_json()
    teacher_id = data.get('teacherId')
    priority = data.get('priority', 1)  # 默认第一志愿
    
    if not teacher_id:
        return jsonify({'success': False, 'message': '请选择导师'}), 400
    
    if priority not in [1, 2]:
        return jsonify({'success': False, 'message': '志愿必须是1或2'}), 400
    
    teacher = User.query.get(teacher_id)
    if not teacher or teacher.role != 'teacher':
        return jsonify({'success': False, 'message': '导师不存在'}), 404
    
    # 检查是否已经选择了该导师
    existing_selection = TeacherSelection.query.filter_by(
        student_id=user_id,
        teacher_id=teacher_id
    ).first()
    if existing_selection:
        return jsonify({'success': False, 'message': '您已经选择过该导师了'}), 400
    
    # 检查是否已经有相同志愿的选择
    existing_priority = TeacherSelection.query.filter_by(
        student_id=user_id,
        priority=priority
    ).first()
    if existing_priority:
        return jsonify({'success': False, 'message': f'您已经设置了第{priority}志愿'}), 400
    
    # 创建选择记录
    selection = TeacherSelection(
        student_id=user_id,
        teacher_id=teacher_id,
        priority=priority,
        status='pending'
    )
    
    try:
        db.session.add(selection)
        db.session.commit()
        
        # 添加申请日志
        log = ApplicationLog(
            user_id=user_id,
            type='double-selection',
            status='pending',
            receipt=f'第{priority}志愿选择导师',
            description=f'第{priority}志愿选择导师：{teacher.realname}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'第{priority}志愿选择导师成功',
            'selection': {
                'id': selection.id,
                'teacherId': teacher_id,
                'teacherName': teacher.realname,
                'priority': priority,
                'status': 'pending',
                'createTime': selection.create_time.strftime('%Y-%m-%d %H:%M:%S')
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'选择失败: {str(e)}'}), 500

# 学生取消选择
@app.route('/api/cancel-selection', methods=['POST'])
@login_required
def cancel_selection():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'student':
        return jsonify({'success': False, 'message': '只有学生可以取消选择'}), 403
    
    # 检查双选阶段
    step = DoubleSelectionStep.query.first()
    if not step or step.current_step != 1:
        return jsonify({'success': False, 'message': '当前不在选导师阶段'}), 400
    
    data = request.get_json()
    selection_id = data.get('selectionId')
    
    if not selection_id:
        return jsonify({'success': False, 'message': '请指定要取消的选择'}), 400
    
    # 检查是否已选择
    selection = TeacherSelection.query.filter_by(
        id=selection_id,
        student_id=user_id
    ).first()
    
    if not selection:
        return jsonify({'success': False, 'message': '选择不存在'}), 404
    
    if selection.status != 'pending':
        return jsonify({'success': False, 'message': '该选择已被处理，无法取消'}), 400
    
    try:
        db.session.delete(selection)
        db.session.commit()
        return jsonify({'success': True, 'message': '取消选择成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'取消选择失败: {str(e)}'}), 500

# 获取选择当前导师的学生列表（导师视角）
@app.route('/api/students-for-teacher', methods=['GET'])
@login_required
def get_students_for_teacher():
    # 获取当前用户（教师）
    token = get_token_from_header()
    payload = decode_token(token)
    user_id = payload.get('user_id')
    
    user = User.query.get(user_id)
    if not user or user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以访问'}), 403
    
    # 获取选择该导师的所有学生
    selections = TeacherSelection.query.filter_by(teacher_id=user_id).all()
    
    students_list = []
    for selection in selections:
        student = User.query.get(selection.student_id)
        if student:
            students_list.append({
                'id': selection.id,
                'studentId': student.username,
                'studentName': student.realname,
                'major': student.major,
                'priority': selection.priority,
                'status': selection.status
            })
    
    return jsonify({'success': True, 'students': students_list})

# 导师确认学生API
@app.route('/api/confirm-selection', methods=['POST'])
@login_required
def confirm_selection():
    data = request.get_json()
    selection_id = data.get('selectionId')
    
    if not selection_id:
        return jsonify({'success': False, 'message': '缺少选择ID'}), 400
    
    # 获取当前用户（教师）
    token = get_token_from_header()
    payload = decode_token(token)
    user_id = payload.get('user_id')
    
    user = db.session.get(User, user_id)
    if not user or user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以操作'}), 403
    
    # 找到选择记录
    selection = TeacherSelection.query.get(selection_id)
    if not selection:
        return jsonify({'success': False, 'message': '选择记录不存在'}), 404
    
    # 检查是否是选择该教师的记录
    if selection.teacher_id != user_id:
        return jsonify({'success': False, 'message': '无权操作此选择'}), 403
    
    # 检查状态
    if selection.status != 'pending':
        return jsonify({'success': False, 'message': '只能确认待确认的选择'}), 400
    
    # 检查教师名额是否已满
    double_teacher = DoubleSelectionTeacher.query.filter_by(teacher_id=user_id).first()
    if not double_teacher:
        return jsonify({'success': False, 'message': '您未参与双选'}), 400
    
    max_quota = double_teacher.max_quota or 5
    if double_teacher.current_quota >= max_quota:
        return jsonify({'success': False, 'message': '您的名额已满'}), 400
    
    # 确认选择
    selection.status = 'confirmed'
    double_teacher.current_quota = (double_teacher.current_quota or 0) + 1
    
    # 注意：不再自动拒绝该学生的其他选择
    # 自动拒绝逻辑将在双选完成阶段统一处理
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '确认学生成功'})

# 导师拒绝学生API
@app.route('/api/reject-selection', methods=['POST'])
@login_required
def reject_selection():
    data = request.get_json()
    selection_id = data.get('selectionId')
    
    if not selection_id:
        return jsonify({'success': False, 'message': '缺少选择ID'}), 400
    
    # 获取当前用户（教师）
    token = get_token_from_header()
    payload = decode_token(token)
    user_id = payload.get('user_id')
    
    user = User.query.get(user_id)
    if not user or user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以操作'}), 403
    
    # 找到选择记录
    selection = TeacherSelection.query.get(selection_id)
    if not selection:
        return jsonify({'success': False, 'message': '选择记录不存在'}), 404
    
    # 检查是否是选择该教师的记录
    if selection.teacher_id != user_id:
        return jsonify({'success': False, 'message': '无权操作此选择'}), 403
    
    # 检查状态
    if selection.status != 'pending':
        return jsonify({'success': False, 'message': '只能拒绝待确认的选择'}), 400
    
    # 拒绝选择
    selection.status = 'rejected'
    # 恢复导师名额
    double_teacher = DoubleSelectionTeacher.query.filter_by(teacher_id=user_id).first()
    if double_teacher:
        double_teacher.current_quota = max(0, double_teacher.current_quota - 1)
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '拒绝学生成功'})

# 导师撤回选择API
@app.route('/api/revoke-selection', methods=['POST'])
@login_required
def revoke_selection():
    data = request.get_json()
    selection_id = data.get('selectionId')
    
    if not selection_id:
        return jsonify({'success': False, 'message': '缺少选择ID'}), 400
    
    # 获取当前用户（教师）
    token = get_token_from_header()
    payload = decode_token(token)
    user_id = payload.get('user_id')
    
    user = User.query.get(user_id)
    if not user or user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以操作'}), 403
    
    # 找到选择记录
    selection = TeacherSelection.query.get(selection_id)
    if not selection:
        return jsonify({'success': False, 'message': '选择记录不存在'}), 404
    
    # 检查是否是选择该教师的记录
    if selection.teacher_id != user_id:
        return jsonify({'success': False, 'message': '无权操作此选择'}), 403
    
    # 检查状态
    if selection.status != 'confirmed' and selection.status != 'rejected':
        return jsonify({'success': False, 'message': '只能撤回已确认或已拒绝的选择'}), 400
    
    # 撤回选择
    selection.status = 'pending'
    # 恢复导师名额
    double_teacher = DoubleSelectionTeacher.query.filter_by(teacher_id=user_id).first()
    if double_teacher:
        double_teacher.current_quota = max(0, double_teacher.current_quota - 1)
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '撤回选择成功'})

# 处理双选结果API（仅管理员）
@app.route('/api/process-selection-results', methods=['POST'])
@login_required
def process_selection_results():
    # 从请求上下文中获取用户信息（JWT方式）
    token = get_token_from_header()
    payload = decode_token(token)
    user_id = payload.get('user_id')
    user = User.query.get(user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以处理双选结果'}), 403
    
    try:
        # 1. 处理第一志愿和第二志愿同时通过的情况
        # 获取所有已确认的选择
        confirmed_selections = TeacherSelection.query.filter_by(status='confirmed').all()
        
        # 按学生分组
        student_selections = {}
        for selection in confirmed_selections:
            if selection.student_id not in student_selections:
                student_selections[selection.student_id] = []
            student_selections[selection.student_id].append(selection)
        
        # 处理多导师确认的情况
        for student_id, selections in student_selections.items():
            if len(selections) > 1:
                # 按优先级排序，选择第一志愿
                selections.sort(key=lambda x: x.priority)
                first_choice = selections[0]
                
                # 拒绝其他选择
                for selection in selections[1:]:
                    selection.status = 'rejected'
        
        # 2. 处理第三阶段仍为pending的学生
        pending_selections = TeacherSelection.query.filter_by(status='pending').all()
        for selection in pending_selections:
            selection.status = 'rejected'
        
        # 3. 处理最终阶段仍为rejected的学生，随机分配
        # 获取参与双选的学生（从double_selection_student表中）
        double_students = DoubleSelectionStudent.query.filter_by(status='active').all()
        students = [ds.student for ds in double_students]
        
        # 获取参与双选的教师
        double_teachers = DoubleSelectionTeacher.query.filter_by(status='active').all()
        teachers = []
        for dt in double_teachers:
            teacher = dt.teacher
            # 为教师添加current_quota和max_quota属性
            teacher.current_quota = dt.current_quota or 0
            teacher.max_quota = dt.max_quota or 5
            teachers.append(teacher)
        
        # 找出所有需要分配的学生（状态为rejected或未被分配的学生）
        rejected_students = []
        for student in students:
            # 检查学生是否有已确认的选择
            has_confirmed = TeacherSelection.query.filter_by(
                student_id=student.id,
                status='confirmed'
            ).first()
            
            # 检查学生是否已经在双选结果表中
            has_result = DoubleSelectionResult.query.filter_by(
                student_id=student.id
            ).first()
            
            # 如果学生没有已确认的选择且不在双选结果表中，需要分配
            if not has_confirmed and not has_result:
                rejected_students.append(student)
        
        # 按当前名额排序教师（从少到多）
        teachers.sort(key=lambda x: x.current_quota or 0)
        
        # 随机分配
        import random
        random.shuffle(rejected_students)
        
        for student in rejected_students:
            # 找到还有名额的教师
            available_teachers = [t for t in teachers if (t.current_quota or 0) < (t.max_quota or 5)]
            if available_teachers:
                # 从名额最少的教师开始选择
                available_teachers.sort(key=lambda x: x.current_quota or 0)
                teacher = available_teachers[0]
                
                # 分配学生
                teacher.current_quota = (teacher.current_quota or 0) + 1
                
                # 创建双选结果
                result = DoubleSelectionResult(
                    student_id=student.id,
                    teacher_id=teacher.id,
                    status='confirmed'
                )
                db.session.add(result)
        
        # 4. 将所有已确认的选择存入双选结果表
        for selection in TeacherSelection.query.filter_by(status='confirmed').all():
            # 检查是否已存在
            existing_result = DoubleSelectionResult.query.filter_by(student_id=selection.student_id).first()
            if not existing_result:
                result = DoubleSelectionResult(
                    student_id=selection.student_id,
                    teacher_id=selection.teacher_id,
                    status='confirmed'
                )
                db.session.add(result)
        
        # 5. 清空teacher_selection表
        TeacherSelection.query.delete()
        
        # 6. 重置参与双选的教师的当前名额
        double_teachers = DoubleSelectionTeacher.query.filter_by(status='active').all()
        for dt in double_teachers:
            dt.current_quota = 0
        
        # 7. 清空除了双选结果表之外的相关表
        DoubleSelectionTeacher.query.delete()
        DoubleSelectionStudent.query.delete()
        DoubleSelectionTime.query.delete()

        db.session.commit()
        
        return jsonify({'success': True, 'message': '双选结果处理成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'处理失败: {str(e)}'}), 500

# 重置双选数据API（仅管理员）
@app.route('/api/reset-double-selection', methods=['POST'])
@login_required
def reset_double_selection():
    # 从请求上下文中获取用户信息（JWT方式）
    token = get_token_from_header()
    payload = decode_token(token)
    user_id = payload.get('user_id')
    user = User.query.get(user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以重置双选数据'}), 403
    
    try:
        # 1. 清空双选结果表
        DoubleSelectionResult.query.delete()
        
        # 2. 清空教师选择表
        TeacherSelection.query.delete()
        
        # 3. 清空双选导师表
        DoubleSelectionTeacher.query.delete()
        
        # 4. 清空双选学生表
        DoubleSelectionStudent.query.delete()
        
        # 5. 清空双选时间表
        DoubleSelectionTime.query.delete()
        
        # 5. 重置所有教师的当前名额
        teachers = User.query.filter_by(role='teacher').all()
        for teacher in teachers:
            teacher.current_quota = 0
        
        # 6. 重置双选阶段为未开始
        step = DoubleSelectionStep.query.first()
        if step:
            step.current_step = 0
        else:
            step = DoubleSelectionStep(current_step=0)
            db.session.add(step)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': '双选数据重置成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'重置失败: {str(e)}'}), 500

# 获取双选结果API
@app.route('/api/double-selection-result', methods=['GET'])
@login_required
def get_double_selection_result():
    # 从请求上下文中获取用户信息（通过@login_required装饰器）
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    try:
        if user.role == 'student':
            # 学生获取自己的双选结果
            result = DoubleSelectionResult.query.filter_by(student_id=user_id).first()
            if result:
                teacher = User.query.get(result.teacher_id)
                return jsonify({
                    'success': True,
                    'result': {
                        'studentId': user_id,
                        'studentName': user.realname,
                        'teacherId': result.teacher_id,
                        'teacherName': teacher.realname if teacher else '未知导师',
                        'status': result.status,
                        'createTime': result.create_time.strftime('%Y-%m-%d %H:%M:%S')
                    }
                })
            else:
                return jsonify({'success': False, 'message': '暂无双选结果'})
        elif user.role == 'teacher':
            # 教师获取分配给自己的学生
            results = DoubleSelectionResult.query.filter_by(teacher_id=user_id).all()
            students = []
            for result in results:
                student = User.query.get(result.student_id)
                if student:
                    students.append({
                        'studentId': student.id,
                        'studentName': student.realname,
                        'studentUsername': student.username,
                        'status': result.status,
                        'createTime': result.create_time.strftime('%Y-%m-%d %H:%M:%S')
                    })
            return jsonify({
                'success': True,
                'students': students
            })
        elif user.role == 'admin':
            # 管理员获取所有双选结果
            results = DoubleSelectionResult.query.all()
            all_results = []
            for result in results:
                student = User.query.get(result.student_id)
                teacher = User.query.get(result.teacher_id)
                if student and teacher:
                    all_results.append({
                        'studentId': student.id,
                        'studentName': student.realname,
                        'studentUsername': student.username,
                        'teacherId': teacher.id,
                        'teacherName': teacher.realname,
                        'status': result.status,
                        'createTime': result.create_time.strftime('%Y-%m-%d %H:%M:%S')
                    })
            return jsonify({
                'success': True,
                'results': all_results
            })
        else:
            return jsonify({'success': False, 'message': '无权访问'}), 403
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取双选结果失败: {str(e)}'}), 500

# ==================== 科研训练API ====================


# 公告管理API

# 获取公告列表
@app.route('/api/announcements', methods=['GET'])
@login_required
def get_announcements():
    try:
        announcements = Announcement.query.order_by(Announcement.create_time.desc()).all()
        
        announcement_list = []
        for announcement in announcements:
            attachments = []
            for attachment in announcement.attachments:
                attachments.append({
                    'id': attachment.id,
                    'filename': attachment.filename,
                    'size': attachment.size
                })
            
            announcement_list.append({
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'author_id': announcement.author_id,
                'author_name': announcement.author.realname,
                'created_at': announcement.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                'attachments': attachments
            })
        
        return jsonify({'success': True, 'announcements': announcement_list})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取公告列表失败: {str(e)}'}), 500

# 获取公告详情
@app.route('/api/announcements/<int:announcement_id>', methods=['GET'])
@login_required
def get_announcement_detail(announcement_id):
    try:
        announcement = Announcement.query.get(announcement_id)
        if not announcement:
            return jsonify({'success': False, 'message': '公告不存在'}), 404
        
        attachments = []
        for attachment in announcement.attachments:
            attachments.append({
                'id': attachment.id,
                'filename': attachment.filename,
                'size': attachment.size
            })
        
        # 标记公告为已读
        user_id = request.current_user['user_id']
        existing_record = AnnouncementReadRecord.query.filter_by(
            user_id=user_id,
            announcement_id=announcement_id
        ).first()
        if not existing_record:
            read_record = AnnouncementReadRecord(
                user_id=user_id,
                announcement_id=announcement_id
            )
            db.session.add(read_record)
            db.session.commit()
        
        return jsonify({
            'success': True,
            'announcement': {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'author_id': announcement.author_id,
                'author_name': announcement.author.realname,
                'created_at': announcement.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                'attachments': attachments
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取公告详情失败: {str(e)}'}), 500

@app.route('/api/announcements/unread-count', methods=['GET'])
@login_required
def get_unread_announcements_count():
    user_id = request.current_user['user_id']
    
    # 获取所有公告
    all_announcements = Announcement.query.all()
    total_announcements = len(all_announcements)
    
    # 获取已读公告数量
    read_records = AnnouncementReadRecord.query.filter_by(user_id=user_id).all()
    read_announcement_ids = [record.announcement_id for record in read_records]
    unread_count = total_announcements - len(read_announcement_ids)
    
    return jsonify({
        'success': True,
        'unread_count': unread_count
    })

@app.route('/api/announcements/mark-all-read', methods=['POST'])
@login_required
def mark_all_announcements_read():
    user_id = request.current_user['user_id']
    
    # 获取所有公告
    all_announcements = Announcement.query.all()
    
    # 为每个公告创建阅读记录（如果不存在）
    for announcement in all_announcements:
        existing_record = AnnouncementReadRecord.query.filter_by(
            user_id=user_id,
            announcement_id=announcement.id
        ).first()
        if not existing_record:
            read_record = AnnouncementReadRecord(
                user_id=user_id,
                announcement_id=announcement.id
            )
            db.session.add(read_record)
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': '所有公告已标记为已读'
    })

# 发布公告（支持附件）
@app.route('/api/announcements', methods=['POST'])
@login_required
def create_announcement():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以发布公告'}), 403
    
    try:
        title = request.form.get('title')
        content = request.form.get('content')
        
        if not title or not content:
            return jsonify({'success': False, 'message': '公告标题和内容不能为空'}), 400
        
        # 创建公告
        announcement = Announcement(
            title=title,
            content=content,
            author_id=user_id
        )
        db.session.add(announcement)
        db.session.flush()  # 获取公告ID
        
        # 处理附件
        # 尝试不同的文件字段名
        files = []
        if 'attachments' in request.files:
            files = request.files.getlist('attachments')
        elif 'attachments[0]' in request.files:
            # 处理前端传递的数组格式
            i = 0
            while f'attachments[{i}]' in request.files:
                files.append(request.files[f'attachments[{i}]'])
                i += 1
        
        for file in files:
            if file and file.filename:
                # 保存文件
                upload_folder = UPLOAD_FOLDER_ATTACHMENTS
                
                # 使用公告标题和原始文件名命名，处理特殊字符
                safe_title = ''.join(c for c in title if c.isalnum() or c in ' _-')
                safe_title = safe_title[:50]  # 限制标题长度
                filename = f"{safe_title}_{file.filename}"
                filepath = os.path.join(upload_folder, filename)
                file.save(filepath)
                
                # 创建附件记录
                attachment = Attachment(
                    filename=file.filename,
                    filepath=filepath,
                    size=os.path.getsize(filepath) // 1024,  # 转换为KB
                    announcement_id=announcement.id
                )
                db.session.add(attachment)
        
        db.session.commit()
        return jsonify({'success': True, 'message': '公告发布成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'发布公告失败: {str(e)}'}), 500

# 编辑公告
@app.route('/api/announcements/<int:announcement_id>', methods=['PUT'])
@login_required
def update_announcement(announcement_id):
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以编辑公告'}), 403
    
    try:
        announcement = Announcement.query.get(announcement_id)
        if not announcement:
            return jsonify({'success': False, 'message': '公告不存在'}), 404
        
        title = request.form.get('title')
        content = request.form.get('content')
        
        if not title or not content:
            return jsonify({'success': False, 'message': '公告标题和内容不能为空'}), 400
        
        # 更新公告内容
        announcement.title = title
        announcement.content = content
        
        # 处理附件删除
        # 检查是否有附件需要删除
        delete_attachments = []
        if 'delete_attachments[0]' in request.form:
            i = 0
            while f'delete_attachments[{i}]' in request.form:
                delete_attachments.append(int(request.form.get(f'delete_attachments[{i}]')))
                i += 1
        
        # 删除指定的附件
        if delete_attachments:
            for attachment_id in delete_attachments:
                attachment = Attachment.query.get(attachment_id)
                if attachment and attachment.announcement_id == announcement.id:
                    # 删除文件
                    import os
                    if os.path.exists(attachment.filepath):
                        os.remove(attachment.filepath)
                    # 删除记录
                    db.session.delete(attachment)
        
        # 处理新附件上传
        # 尝试不同的文件字段名
        files = []
        if 'attachments' in request.files:
            files = request.files.getlist('attachments')
        elif 'attachments[0]' in request.files:
            # 处理前端传递的数组格式
            i = 0
            while f'attachments[{i}]' in request.files:
                files.append(request.files[f'attachments[{i}]'])
                i += 1
        
        # 确保即使只有一个文件也能正确处理
        if 'attachments' in request.files and len(request.files.getlist('attachments')) > 0:
            files = request.files.getlist('attachments')
        
        # 添加新附件
        if files:
            for file in files:
                if file and file.filename:
                    # 保存文件
                    upload_folder = UPLOAD_FOLDER_ATTACHMENTS
                    
                    # 使用公告标题和原始文件名命名，处理特殊字符
                    safe_title = ''.join(c for c in title if c.isalnum() or c in ' _-')
                    safe_title = safe_title[:50]  # 限制标题长度
                    filename = f"{safe_title}_{file.filename}"
                    filepath = os.path.join(upload_folder, filename)
                    file.save(filepath)
                    
                    # 创建附件记录
                    attachment = Attachment(
                        filename=file.filename,
                        filepath=filepath,
                        size=os.path.getsize(filepath) // 1024,  # 转换为KB
                        announcement_id=announcement.id
                    )
                    db.session.add(attachment)
        
        db.session.commit()
        return jsonify({'success': True, 'message': '公告编辑成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'编辑公告失败: {str(e)}'}), 500

# 删除公告
@app.route('/api/announcements/<int:announcement_id>', methods=['DELETE'])
@login_required
def delete_announcement(announcement_id):
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以删除公告'}), 403
    
    try:
        announcement = Announcement.query.get(announcement_id)
        if not announcement:
            return jsonify({'success': False, 'message': '公告不存在'}), 404
        
        # 删除附件文件
        for attachment in announcement.attachments:
            import os
            if os.path.exists(attachment.filepath):
                os.remove(attachment.filepath)
        
        # 删除公告
        db.session.delete(announcement)
        db.session.commit()
        return jsonify({'success': True, 'message': '公告删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'删除公告失败: {str(e)}'}), 500

# 下载附件
@app.route('/api/attachments/<int:attachment_id>', methods=['GET'])
@login_required
def download_attachment(attachment_id):
    try:
        attachment = Attachment.query.get(attachment_id)
        if not attachment:
            return jsonify({'success': False, 'message': '附件不存在'}), 404
        
        import os
        if not os.path.exists(attachment.filepath):
            return jsonify({'success': False, 'message': '附件文件不存在'}), 404
        
        from flask import send_file
        return send_file(attachment.filepath, as_attachment=True, download_name=attachment.filename)
    except Exception as e:
        return jsonify({'success': False, 'message': f'下载附件失败: {str(e)}'}), 500

# 科研训练项目报名
@app.route('/api/research-projects/register', methods=['POST'])
@login_required
def register_research_project():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'student':
        return jsonify({'success': False, 'message': '只有学生可以报名科研训练项目'}), 403
    
    try:
        data = request.get_json()
        project_id = data.get('projectId')
        
        if not project_id:
            return jsonify({'success': False, 'message': '项目ID不能为空'}), 400
        
        # 检查项目是否存在
        project = ResearchProject.query.get(project_id)
        if not project:
            return jsonify({'success': False, 'message': '项目不存在'}), 404
        
        # 检查是否已经报名
        existing_registration = ResearchRegistration.query.filter_by(
            student_id=user_id, project_id=project_id
        ).first()
        if existing_registration:
            return jsonify({'success': False, 'message': '您已经报名了该项目'}), 400
        
        # 创建报名记录
        registration = ResearchRegistration(
            student_id=user_id,
            project_id=project_id,
            status='pending'
        )
        db.session.add(registration)
        
        # 更新项目的已报名人数
        project.registered_count = project.registered_count + 1 if project.registered_count else 1
        
        db.session.commit()
        
        # 添加申请日志
        log = ApplicationLog(
            user_id=user_id,
            type='train',
            status='pending',
            receipt='报名科研训练项目',
            description=f'报名科研训练项目：{project.project_name}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '报名成功'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'报名失败: {str(e)}'}), 500

# 获取教师的待处理报名请求
@app.route('/api/research-projects/registrations/pending', methods=['GET'])
@login_required
def get_pending_registrations():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以查看报名请求'}), 403
    
    try:
        # 获取教师的项目
        projects = ResearchProject.query.filter_by(teacher_id=user_id).all()
        project_ids = [p.id for p in projects]
        
        # 获取这些项目的待处理报名
        registrations = ResearchRegistration.query.filter(
            ResearchRegistration.project_id.in_(project_ids),
            ResearchRegistration.status == 'pending'
        ).all()
        
        registration_list = []
        for registration in registrations:
            project = registration.project
            student = registration.student
            if project and student:
                registration_list.append({
                    'id': registration.id,
                    'projectId': project.id,
                    'projectName': project.project_name,
                    'studentId': student.id,
                    'studentName': student.realname,
                    'studentUsername': student.username,
                    'studentMajor': student.major,
                    'studentDepartment': student.department,
                    'studentContact': student.contact,
                    'studentEmail': student.email,
                    'status': registration.status,
                    'createTime': registration.create_time.strftime('%Y-%m-%d %H:%M:%S')
                })
        
        return jsonify({
            'success': True,
            'registrations': registration_list,
            'projectCount': len(project_ids),
            'registrationCount': len(registration_list)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取报名请求失败: {str(e)}'}), 500

# 处理报名请求
@app.route('/api/research-projects/registrations/<int:registration_id>', methods=['PUT'])
@login_required
def handle_registration(registration_id):
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以处理报名请求'}), 403
    
    try:
        data = request.get_json()
        registration_id = data.get('registrationId')
        action = data.get('action')  # 'approve' or 'reject'
        
        if not registration_id or not action:
            return jsonify({'success': False, 'message': '参数不能为空'}), 400
        
        registration = ResearchRegistration.query.get(registration_id)
        if not registration:
            return jsonify({'success': False, 'message': '报名记录不存在'}), 404
        
        # 检查是否是该项目的导师
        project = registration.project
        if project.teacher_id != user_id:
            return jsonify({'success': False, 'message': '您不是该项目的导师'}), 403
        
        if action == 'approve':
            registration.status = 'confirmed'
            project.confirmed_count = project.confirmed_count + 1 if project.confirmed_count else 1
        elif action == 'reject':
            registration.status = 'rejected'
            project.registered_count = project.registered_count - 1 if project.registered_count else 0
        else:
            return jsonify({'success': False, 'message': '无效的操作'}), 400
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'报名请求{action == "approve" and "已批准" or "已拒绝"}'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'处理失败: {str(e)}'}), 500

# 获取已确认的科研训练项目(学生与老师)
@app.route('/api/research-projects/confirmed', methods=['GET'])
@login_required
def get_confirmed_research_projects():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'student' and user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有学生或老师可以查看已确认的项目'}), 403
    
    try:
        # 获取已确认的项目
        registrations = ResearchRegistration.query.filter_by(
            student_id=user_id, status='confirmed'
        ).all()
        
        confirmed_projects = []
        for registration in registrations:
            project = registration.project
            if project:
                confirmed_projects.append({
                    'id': project.id,
                    'batch': project.batch,
                    'projectName': project.project_name,
                    'direction': project.direction,
                    'teacherId': project.teacher_id,
                    'teacherName': project.teacher.realname if project.teacher else '未知',
                    'teacherUsername': project.teacher.username if project.teacher else '未知',
                    'teacherDepartment': project.teacher.department if project.teacher and hasattr(project.teacher, 'department') else '未知',
                    'teacherContact': project.teacher.contact if project.teacher and hasattr(project.teacher, 'contact') else '未知',
                    'teacherTitle': project.teacher.title if project.teacher and hasattr(project.teacher, 'title') else '未知',
                    'teacherEmail': project.teacher.email if project.teacher and hasattr(project.teacher, 'email') else '未知',
                    'major': project.major,
                    'department': project.department,
                    'isNational': project.is_national,
                    'nationalLab': project.national_lab,
                    'maxStudents': project.max_students,
                    'registeredCount': project.registered_count,
                    'confirmedCount': project.confirmed_count,
                    'status': registration.status,
                    'createTime': registration.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'registered': True
                })
        
        return jsonify({
            'success': True,
            'projects': confirmed_projects
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取已确认项目失败: {str(e)}'}), 500


@app.route('/api/research-projects/teacher', methods=['GET'])
@login_required
def get_teacher_projects():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以查看自己的课题'}), 403
    
    try:
        # 获取教师的课题，状态为approved
        projects = ResearchProject.query.filter_by(
            teacher_id=user_id, status='approved'
        ).all()
        
        project_list = []
        for project in projects:
            project_list.append({
                'id': project.id,
                'batch': project.batch,
                'projectName': project.project_name,
                'direction': project.direction,
                'teacherId': project.teacher_id,
                'teacherName': project.teacher.realname if project.teacher else '未知',
                'teacherUsername': project.teacher.username if project.teacher else '未知',
                'teacherDepartment': project.teacher.department if project.teacher and hasattr(project.teacher, 'department') else '未知',
                'teacherContact': project.teacher.contact if project.teacher and hasattr(project.teacher, 'contact') else '未知',
                'teacherTitle': project.teacher.title if project.teacher and hasattr(project.teacher, 'title') else '未知',
                'teacherEmail': project.teacher.email if project.teacher and hasattr(project.teacher, 'email') else '未知',
                'major': project.major,
                'department': project.department,
                'isNational': project.is_national,
                'nationalLab': project.national_lab,
                'maxStudents': project.max_students,
                'registeredCount': project.registered_count,
                'confirmedCount': project.confirmed_count,
                'status': project.status,
                'createTime': project.create_time.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({
            'success': True,
            'projects': project_list
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取教师课题失败: {str(e)}'}), 500

# 获取学生的报名记录
@app.route('/api/research-projects/registrations', methods=['GET'])
@login_required
def get_student_registrations():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'student':
        return jsonify({'success': False, 'message': '只有学生可以查看报名记录'}), 403
    
    try:
        # 获取学生的报名记录
        registrations = ResearchRegistration.query.filter_by(student_id=user_id).all()
        
        registration_projects = []
        for registration in registrations:
            project = registration.project
            if project:
                registration_projects.append({
                    'id': project.id,
                    'batch': project.batch,
                    'projectName': project.project_name,
                    'direction': project.direction,
                    'teacherId': project.teacher_id,
                    'teacherName': project.teacher.realname if project.teacher else '未知',
                    'teacherUsername': project.teacher.username if project.teacher else '未知',
                    'teacherDepartment': project.teacher.department if project.teacher and hasattr(project.teacher, 'department') else '未知',
                    'teacherContact': project.teacher.contact if project.teacher and hasattr(project.teacher, 'contact') else '未知',
                    'teacherTitle': project.teacher.title if project.teacher and hasattr(project.teacher, 'title') else '未知',
                    'teacherEmail': project.teacher.email if project.teacher and hasattr(project.teacher, 'email') else '未知',
                    'major': project.major,
                    'department': project.department,
                    'isNational': project.is_national,
                    'nationalLab': project.national_lab,
                    'maxStudents': project.max_students,
                    'registeredCount': project.registered_count,
                    'confirmedCount': project.confirmed_count,
                    'status': registration.status,
                    'createTime': registration.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'registered': True
                })
        
        return jsonify({
            'success': True,
            'projects': registration_projects
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取报名记录失败: {str(e)}'}), 500

# 查看科研训练项目详情
@app.route('/api/research-projects/<int:project_id>', methods=['GET'])
@login_required
def get_research_project_detail(project_id):
    try:
        project = ResearchProject.query.get(project_id)
        if not project:
            return jsonify({'success': False, 'message': '项目不存在'}), 404
        
        return jsonify({
            'success': True,
            'project': {
                'id': project.id,
                'batch': project.batch,
                'projectName': project.project_name,
                'direction': project.direction,
                'teacherId': project.teacher_id,
                'teacherName': project.teacher.realname if project.teacher else '未知',
                'teacherUsername': project.teacher.username if project.teacher else '未知',
                'teacherDepartment': project.teacher.department if project.teacher and hasattr(project.teacher, 'department') else '未知',
                'teacherContact': project.teacher.contact if project.teacher and hasattr(project.teacher, 'contact') else '未知',
                'teacherTitle': project.teacher.title if project.teacher and hasattr(project.teacher, 'title') else '未知',
                'teacherEmail': project.teacher.email if project.teacher and hasattr(project.teacher, 'email') else '未知',
                'maxStudents': project.max_students,
                'registeredCount': project.registered_count,
                'confirmedCount': project.confirmed_count,
                'major': project.major,
                'department': project.department,
                'isNational': project.is_national,
                'nationalLab': project.national_lab,
                'description': project.description,
                'requirements': project.requirements
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取项目详情失败: {str(e)}'}), 500

# 教师申请科研训练项目
@app.route('/api/research-projects/apply', methods=['POST'])
@login_required
def apply_research_project():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'teacher':
        return jsonify({'success': False, 'message': '只有教师可以申请科研训练项目'}), 403
    
    try:
        data = request.get_json()
        
        # 验证必填字段
        required_fields = ['batch', 'projectName', 'direction', 'maxStudents', 'major', 'department']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field}不能为空'}), 400
        
        # 创建项目
        project = ResearchProject(
            batch=data['batch'],
            project_name=data['projectName'],
            direction=data['direction'],
            teacher_id=user_id,
            max_students=data['maxStudents'],
            major=data['major'],
            department=data['department'],
            is_national=data.get('isNational', False),
            national_lab=data.get('nationalLab', ''),
            description=data.get('description', ''),
            requirements=data.get('requirements', ''),
            status='pending'  # 初始状态为待审核
        )
        
        db.session.add(project)
        db.session.commit()
        
        # 添加申请日志
        log = ApplicationLog(
            user_id=user_id,
            type='project',
            status='pending',
            receipt='科研训练项目申请',
            description=f'申请科研训练项目：{data["projectName"]}'
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '项目申请提交成功，等待管理员审核',
            'projectId': project.id
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'提交申请失败: {str(e)}'}), 500

# 管理员获取待审核的科研训练项目
@app.route('/api/research-projects/pending', methods=['GET'])
@login_required
def get_pending_research_projects():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以查看待审核项目'}), 403
    
    try:
        # 获取待审核的项目
        projects = ResearchProject.query.filter_by(status='pending').all()
        
        pending_projects = []
        for project in projects:
            pending_projects.append({
                'id': project.id,
                'projectName': project.project_name,
                'direction': project.direction,
                'teacherName': project.teacher.realname if project.teacher else '未知',
                'teacherUsername': project.teacher.username if project.teacher else '未知',
                'teacherDepartment': project.teacher.department if project.teacher and hasattr(project.teacher, 'department') else '未知',
                'teacherContact': project.teacher.contact if project.teacher and hasattr(project.teacher, 'contact') else '未知',
                'teacherTitle': project.teacher.title if project.teacher and hasattr(project.teacher, 'title') else '未知',
                'teacherEmail': project.teacher.email if project.teacher and hasattr(project.teacher, 'email') else '未知',
                'batch': project.batch,
                'status': project.status,
                'major': project.major,
                'department': project.department,
                'isNational': project.is_national,
                'nationalLab': project.national_lab,
                'maxStudents': project.max_students,
                'registeredCount': project.registered_count,
                'confirmedCount': project.confirmed_count
            })
        
        return jsonify({
            'success': True,
            'projects': pending_projects
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取待审核项目失败: {str(e)}'}), 500

# 管理员批准科研训练项目
@app.route('/api/research-projects/approve', methods=['POST'])
@login_required
def approve_research_project():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以批准项目'}), 403
    
    try:
        data = request.get_json()
        project_id = data.get('projectId')
        
        if not project_id:
            return jsonify({'success': False, 'message': '项目ID不能为空'}), 400
        
        # 查找项目
        project = ResearchProject.query.get(project_id)
        if not project:
            return jsonify({'success': False, 'message': '项目不存在'}), 404
        
        if project.status != 'pending':
            return jsonify({'success': False, 'message': '项目状态不是待审核'}), 400
        
        # 更新状态为已批准
        project.status = 'approved'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '项目批准成功'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'批准项目失败: {str(e)}'}), 500

# 管理员拒绝科研训练项目
@app.route('/api/research-projects/reject', methods=['POST'])
@login_required
def reject_research_project():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    if user.role != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以拒绝项目'}), 403
    
    try:
        data = request.get_json()
        project_id = data.get('projectId')
        
        if not project_id:
            return jsonify({'success': False, 'message': '项目ID不能为空'}), 400
        
        # 查找项目
        project = ResearchProject.query.get(project_id)
        if not project:
            return jsonify({'success': False, 'message': '项目不存在'}), 404
        
        if project.status != 'pending':
            return jsonify({'success': False, 'message': '项目状态不是待审核'}), 400
        
        # 更新状态为已拒绝
        project.status = 'rejected'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '项目拒绝成功'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'拒绝项目失败: {str(e)}'}), 500

# 更新获取科研训练项目列表API，只返回已批准的项目
@app.route('/api/research-projects', methods=['GET'])
@login_required
def get_research_projects():
    user_id = request.current_user['user_id']
    user = User.query.get(user_id)
    
    try:
        # 获取查询参数
        batch = request.args.get('batch', '')
        direction = request.args.get('direction', '')
        project_name = request.args.get('projectName', '')
        teacher_username = request.args.get('teacherUsername', '')
        teacher_name = request.args.get('teacherName', '')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('pageSize', 10))
        all = request.args.get('all', 'false').lower() == 'true'
        
        # 构建查询，只返回已批准的项目
        query = ResearchProject.query.filter_by(status='approved')
        
        # 应用筛选条件
        if batch:
            query = query.filter(ResearchProject.batch == batch)
        if direction:
            query = query.filter(ResearchProject.direction.like(f'%{direction}%'))
        if project_name:
            query = query.filter(ResearchProject.project_name.like(f'%{project_name}%'))
        if teacher_username:
            query = query.join(User).filter(User.username.like(f'%{teacher_username}%'))
        if teacher_name:
            query = query.join(User).filter(User.realname.like(f'%{teacher_name}%'))
        
        # 计算总记录数
        total = query.count()
        
        # 分页
        if all:
            projects = query.all()
        else:
            projects = query.offset((page - 1) * page_size).limit(page_size).all()
        
        # 构建返回数据
        project_list = []
        for project in projects:
            project_list.append({
                'id': project.id,
                'batch': project.batch,
                'projectName': project.project_name,
                'direction': project.direction,
                'teacherId': project.teacher_id,
                'teacherName': project.teacher.realname if project.teacher else '未知',
                'teacherUsername': project.teacher.username if project.teacher else '未知',
                'teacherDepartment': project.teacher.department if project.teacher and hasattr(project.teacher, 'department') else '未知',
                'teacherContact': project.teacher.contact if project.teacher and hasattr(project.teacher, 'contact') else '未知',
                'teacherTitle': project.teacher.title if project.teacher and hasattr(project.teacher, 'title') else '未知',
                'teacherEmail': project.teacher.email if project.teacher and hasattr(project.teacher, 'email') else '未知',
                'maxStudents': project.max_students,
                'registeredCount': project.registered_count,
                'confirmedCount': project.confirmed_count,
                'major': project.major,
                'department': project.department,
                'isNational': project.is_national,
                'nationalLab': project.national_lab
            })
        
        return jsonify({
            'success': True,
            'projects': project_list,
            'total': total,
            'page': page,
            'pageSize': page_size
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取科研训练项目失败: {str(e)}'}), 500
               
    

# 确保上传目录存在
os.makedirs(UPLOAD_FOLDER_RESUME, exist_ok=True)

# 获取实习列表
@app.route('/api/internship/list', methods=['GET'])
@login_required
def get_internship_list():
    """获取实习列表"""
    try:
        # 获取所有激活状态的实习
        internships = Internship.query.filter_by(status='active').all()
        
        # 构建响应数据
        internship_list = []
        for internship in internships:
            # 计算已报名人数
            registered_count = InternshipApplication.query.filter_by(
                internship_id=internship.id
            ).count()
            
            # 检查当前用户是否已报名
            user_id = request.current_user['user_id']
            has_applied = InternshipApplication.query.filter_by(
                student_id=user_id,
                internship_id=internship.id
            ).first() is not None
            
            internship_list.append({
                'id': internship.id,
                'company': internship.company.company_name,
                'position': internship.position,
                'city': internship.city,
                'location': internship.location,
                'salary': internship.salary or '',
                'skillTags': internship.skill_tags.split(',') if internship.skill_tags else [],
                'welfareTags': internship.welfare_tags.split(',') if internship.welfare_tags else [],
                'quota': internship.quota,
                'registeredCount': registered_count,
                'deadline': internship.deadline.strftime('%Y-%m-%d'),
                'registered': has_applied,
                'title': internship.title,
                'description': internship.description,
                'requirements': internship.requirements
            })
        
        return jsonify({'success': True, 'internships': internship_list})
    except Exception as e:
        print(f"获取实习列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

# 初始化数据库
with app.app_context():
    db.create_all()
    
    # 生成测试用户
    print('开始生成测试用户...')
    
    # 1. 创建管理员账号
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            password_hash=bcrypt.generate_password_hash('admin123').decode('utf-8'),
            realname='管理员',
            role='admin',
            first_login=False
        )
        db.session.add(admin)
        print('管理员账号创建成功: admin/admin123')
    
    # 2. 创建示例学生账号
    student = User.query.filter_by(username='2021001').first()
    if not student:
        student = User(
            username='2021001',
            password_hash=bcrypt.generate_password_hash('123456').decode('utf-8'),
            id_card_last6='123456',
            realname='张三',
            role='student',
            major='计算机科学',
            first_login=True
        )
        db.session.add(student)
        print('学生账号创建成功: 2021001/123456')
    
    # 3. 创建示例教师账号
    teacher = User.query.filter_by(username='T001').first()
    if not teacher:
        teacher = User(
            username='T001',
            password_hash=bcrypt.generate_password_hash('654321').decode('utf-8'),
            id_card_last6='654321',
            realname='李老师',
            role='teacher',
            major='计算机科学',
            min_quota=1,
            max_quota=5,
            current_quota=0,
            first_login=True
        )
        db.session.add(teacher)
        print('教师账号创建成功: T001/654321')
    
    # 4. 创建更多测试教师
    teachers_data = [
        {'username': 'T002', 'realname': '王老师', 'major': '软件工程'},
        {'username': 'T003', 'realname': '张老师', 'major': '人工智能'},
        {'username': 'T004', 'realname': '刘老师', 'major': '网络安全'}
    ]
    
    for t_data in teachers_data:
        teacher = User.query.filter_by(username=t_data['username']).first()
        if not teacher:
            teacher = User(
                username=t_data['username'],
                password_hash=bcrypt.generate_password_hash('123456').decode('utf-8'),
                id_card_last6='123456',
                realname=t_data['realname'],
                role='teacher',
                major=t_data['major'],
                min_quota=1,
                max_quota=5,
                current_quota=0,
                first_login=True
            )
            db.session.add(teacher)
            print(f'教师账号创建成功: {t_data["username"]}/123456')
    
    # 提交所有更改
    db.session.commit()
    print('测试用户生成完成！')

# 实验报告相关API
@app.route('/api/experiment-reports', methods=['GET'])
@login_required
def get_experiment_reports():
    """获取实验报告列表（我的实验报告）"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        
        # 无论角色是谁，都只显示自己的实验报告
        reports = ExperimentReport.query.filter_by(student_id=user_id).order_by(ExperimentReport.create_time.desc()).all()
        
        result = []
        for report in reports:
            project_name = report.project.project_name if report.project else ''
            result.append({
                'id': report.id,
                'student_id': report.student_id,
                'studentName': report.student.realname if report.student else '',
                'studentUsername': report.student.username if report.student else '',
                'project_id': report.project_id,
                'projectName': project_name,
                'title': report.title,
                'content': report.content,
                'date': report.date.strftime('%Y-%m-%d'),
                'fileUrl': report.file_url,
                'createTime': report.create_time.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({'success': True, 'reports': result})
    except Exception as e:
        print(f"获取实验报告列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取实验报告列表失败: {str(e)}'}), 500

@app.route('/api/experiment-reports', methods=['POST'])
@login_required
def upload_experiment_report():
    """上传实验报告"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        
        if user.role != 'student':
            return jsonify({'success': False, 'message': '只有学生可以上传实验报告'}), 403
        
        # 获取表单数据
        title = request.form.get('title')
        content = request.form.get('content')
        date_str = request.form.get('date')
        project_id = request.form.get('project_id')
        
        if not title or not content or not date_str:
            return jsonify({'success': False, 'message': '请填写完整的实验报告信息'}), 400
        
        # 验证项目ID
        if not project_id:
            return jsonify({'success': False, 'message': '请选择科研项目'}), 400
        
        try:
            project_id_int = int(project_id)
        except ValueError:
            return jsonify({'success': False, 'message': '项目ID格式不正确'}), 400
        
        # 验证项目是否存在
        project = ResearchProject.query.get(project_id_int)
        if not project:
            return jsonify({'success': False, 'message': '科研项目不存在'}), 400
        
        # 验证学生是否已报名该项目
        registration = ResearchRegistration.query.filter_by(
            student_id=user_id,
            project_id=project_id_int,
            status='confirmed'
        ).first()
        if not registration:
            return jsonify({'success': False, 'message': '您未报名或未确认该科研项目，无法上传实验报告'}), 400
        
        # 解析日期
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': '日期格式不正确，请使用YYYY-MM-DD格式'}), 400
        
        # 处理文件上传
        file_url = None
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename:
                import uuid
                filename = f"{uuid.uuid4()}_{file.filename}"
                filepath = os.path.join(UPLOAD_FOLDER_EXPERIMENT_REPORTS, filename)
                file.save(filepath)
                file_url = f"/uploads/experiment_reports/{filename}"
        
        # 创建实验报告记录
        new_report = ExperimentReport(
            student_id=user_id,
            project_id=project_id_int,
            title=title,
            content=content,
            date=date,
            file_url=file_url
        )
        
        db.session.add(new_report)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '实验报告上传成功'})
    except Exception as e:
        db.session.rollback()
        print(f"上传实验报告失败: {str(e)}")
        return jsonify({'success': False, 'message': f'上传实验报告失败: {str(e)}'}), 500

@app.route('/api/experiment-reports/<int:report_id>', methods=['PUT'])
@login_required
def update_experiment_report(report_id):
    """编辑实验报告"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        
        # 查找实验报告
        report = ExperimentReport.query.get(report_id)
        if not report:
            return jsonify({'success': False, 'message': '实验报告不存在'}), 404
        
        # 检查权限
        if user.role != 'student' or report.student_id != user_id:
            return jsonify({'success': False, 'message': '无权限编辑此实验报告'}), 403
        
        # 获取表单数据
        title = request.form.get('title')
        content = request.form.get('content')
        date_str = request.form.get('date')
        
        if not title or not content or not date_str:
            return jsonify({'success': False, 'message': '请填写完整的实验报告信息'}), 400
        
        # 解析日期
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': '日期格式不正确，请使用YYYY-MM-DD格式'}), 400
        
        # 处理文件上传
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename:
                # 删除旧文件
                if report.file_url:
                    old_filename = report.file_url.split('/')[-1]
                    old_filepath = os.path.join(UPLOAD_FOLDER_EXPERIMENT_REPORTS, old_filename)
                    if os.path.exists(old_filepath):
                        os.remove(old_filepath)
                
                # 保存新文件
                import uuid
                filename = f"{uuid.uuid4()}_{file.filename}"
                filepath = os.path.join(UPLOAD_FOLDER_EXPERIMENT_REPORTS, filename)
                file.save(filepath)
                report.file_url = f"/uploads/experiment_reports/{filename}"
        
        # 更新实验报告
        report.title = title
        report.content = content
        report.date = date
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': '实验报告更新成功'})
    except Exception as e:
        db.session.rollback()
        print(f"更新实验报告失败: {str(e)}")
        return jsonify({'success': False, 'message': f'更新实验报告失败: {str(e)}'}), 500

@app.route('/api/experiment-reports/<int:report_id>', methods=['DELETE'])
@login_required
def delete_experiment_report(report_id):
    """删除实验报告"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        
        # 查找实验报告
        report = ExperimentReport.query.get(report_id)
        if not report:
            return jsonify({'success': False, 'message': '实验报告不存在'}), 404
        
        # 检查权限
        if user.role != 'student' or report.student_id != user_id:
            return jsonify({'success': False, 'message': '无权限删除此实验报告'}), 403
        
        # 删除文件
        if report.file_url:
            filename = report.file_url.split('/')[-1]
            filepath = os.path.join(UPLOAD_FOLDER_EXPERIMENT_REPORTS, filename)
            if os.path.exists(filepath):
                os.remove(filepath)
        
        # 删除实验报告记录
        db.session.delete(report)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '实验报告删除成功'})
    except Exception as e:
        db.session.rollback()
        print(f"删除实验报告失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除实验报告失败: {str(e)}'}), 500

@app.route('/api/experiment-reports/students', methods=['GET'])
@login_required
def get_student_experiment_reports():
    """获取学生实验报告列表（教师用）"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        
        if user.role != 'teacher' and user.role != 'admin':
            return jsonify({'success': False, 'message': '只有教师和管理员可以查看学生实验报告'}), 403
        
        if user.role == 'teacher':
            # 教师查看自己负责项目下学生的实验报告
            # 获取教师负责的所有科研项目
            projects = ResearchProject.query.filter_by(teacher_id=user_id).all()
            project_ids = [p.id for p in projects]
            
            # 获取这些项目下学生的实验报告
            reports = ExperimentReport.query.filter(
                ExperimentReport.project_id.in_(project_ids)
            ).order_by(ExperimentReport.create_time.desc()).all()
        else:
            # 管理员可以查看所有学生的实验报告
            reports = ExperimentReport.query.order_by(ExperimentReport.create_time.desc()).all()
        
        result = []
        for report in reports:
            result.append({
                'id': report.id,
                'student_id': report.student_id,
                'studentName': report.student.realname if report.student else '',
                'studentUsername': report.student.username if report.student else '',
                'title': report.title,
                'content': report.content,
                'date': report.date.strftime('%Y-%m-%d'),
                'fileUrl': report.file_url,
                'createTime': report.create_time.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({'success': True, 'reports': result})
    except Exception as e:
        print(f"获取学生实验报告列表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取学生实验报告列表失败: {str(e)}'}), 500


@app.route('/api/database/stats', methods=['GET'])
@login_required
def get_database_stats():
    """获取数据库统计信息"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        if user.role != 'admin':
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        # 获取统计数据
        total_users = User.query.count()
        students = User.query.filter_by(role='student').count()
        teachers = User.query.filter_by(role='teacher').count()
        companies = Company.query.count()
        
        return jsonify({
            'success': True,
            'data': {
                'totalUsers': total_users,
                'students': students,
                'teachers': teachers,
                'companies': companies
            }
        })
    except Exception as e:
        print(f"获取数据库统计失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})

@app.route('/api/database/table/<table_name>', methods=['GET'])
@login_required
def get_table_data(table_name):
    """获取指定数据表的数据"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        if user.role != 'admin':
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        # 允许访问的表
        allowed_tables = [
            'user', 'company', 'internship', 'internship_application',
            'team', 'team_member', 'teacher_selection',
            'double_selection_teacher', 'double_selection_student',
            'train_project', 'train_application', 'train_report',
            'application_log'
        ]
        
        if table_name not in allowed_tables:
            return jsonify({'success': False, 'message': '不允许访问该表'}), 403
        
        # 获取分页参数
        page = int(request.args.get('page', 1))
        size = int(request.args.get('size', 20))
        offset = (page - 1) * size
        
        # 根据表名获取数据
        table_class = None
        if table_name == 'user':
            table_class = User
        elif table_name == 'company':
            table_class = Company
        elif table_name == 'internship':
            table_class = Internship
        elif table_name == 'internship_application':
            table_class = InternshipApplication
        elif table_name == 'team':
            table_class = Team
        elif table_name == 'team_member':
            table_class = TeamMember
        elif table_name == 'teacher_selection':
            table_class = TeacherSelection
        elif table_name == 'double_selection_teacher':
            table_class = DoubleSelectionTeacher
        elif table_name == 'double_selection_student':
            table_class = DoubleSelectionStudent
        elif table_name == 'train_project':
            table_class = TrainProject
        elif table_name == 'train_application':
            table_class = TrainApplication
        elif table_name == 'train_report':
            table_class = TrainReport
        elif table_name == 'application_log':
            table_class = ApplicationLog
        
        if not table_class:
            return jsonify({'success': False, 'message': '表不存在'}), 404
        
        # 查询数据
        total_count = table_class.query.count()
        total_pages = (total_count + size - 1) // size
        
        data = table_class.query.offset(offset).limit(size).all()
        
        # 获取列名
        columns = [c.key for c in table_class.__table__.columns]
        
        # 转换为字典列表
        data_list = []
        for row in data:
            row_dict = {}
            for col in columns:
                value = getattr(row, col)
                # 处理特殊类型
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, bytes):
                    value = value.decode('utf-8', errors='ignore')
                row_dict[col] = value
            data_list.append(row_dict)
        
        return jsonify({
            'success': True,
            'columns': columns,
            'data': data_list,
            'total_pages': total_pages,
            'current_page': page
        })
    except Exception as e:
        print(f"获取表数据失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取失败: {str(e)}'})


@app.route('/api/database/table/<table_name>/delete/<int:record_id>', methods=['DELETE'])
@login_required
def delete_table_record(table_name, record_id):
    """删除指定数据表中的记录"""
    try:
        user_id = request.current_user['user_id']
        user = User.query.get(user_id)
        if user.role != 'admin':
            return jsonify({'success': False, 'message': '权限不足'}), 403
        
        # 允许访问的表
        allowed_tables = [
            'user', 'company', 'internship', 'internship_application',
            'team', 'team_member', 'teacher_selection',
            'double_selection_teacher', 'double_selection_student',
            'train_project', 'train_application', 'train_report',
            'application_log'
        ]
        
        if table_name not in allowed_tables:
            return jsonify({'success': False, 'message': '不允许访问该表'}), 403
        
        # 根据表名获取模型类
        table_class = None
        if table_name == 'user':
            table_class = User
        elif table_name == 'company':
            table_class = Company
        elif table_name == 'internship':
            table_class = Internship
        elif table_name == 'internship_application':
            table_class = InternshipApplication
        elif table_name == 'team':
            table_class = Team
        elif table_name == 'team_member':
            table_class = TeamMember
        elif table_name == 'teacher_selection':
            table_class = TeacherSelection
        elif table_name == 'double_selection_teacher':
            table_class = DoubleSelectionTeacher
        elif table_name == 'double_selection_student':
            table_class = DoubleSelectionStudent
        elif table_name == 'train_project':
            table_class = TrainProject
        elif table_name == 'train_application':
            table_class = TrainApplication
        elif table_name == 'train_report':
            table_class = TrainReport
        elif table_name == 'application_log':
            table_class = ApplicationLog
        
        if not table_class:
            return jsonify({'success': False, 'message': '表不存在'}), 404
        
        # 查询并删除记录
        record = table_class.query.get(record_id)
        if not record:
            return jsonify({'success': False, 'message': '记录不存在'}), 404
        
        db.session.delete(record)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        print(f"删除记录失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})
if __name__ == '__main__':
    app.run(debug=True, port=5000, host='0.0.0.0')